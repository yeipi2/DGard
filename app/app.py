import re
import os
import json
import psycopg2
import psycopg2.extras
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_from_directory
from datetime import date, datetime
 
app = Flask(__name__)
app.secret_key = '12345' 


@app.route('/')
def index():
    return render_template('index.html')

@app.route('/panel')
def panel():
    # Verificar si el usuario ha iniciado sesión
    if 'user_email' not in session:
        flash('Debe iniciar sesión para acceder al panel', 'error')
        return redirect(url_for('index'))
    return redirect(url_for('perfil'))

@app.route('/login', methods=['POST'])
def login():
    email = request.form['email'].strip()
    password = request.form['password']
    
    # Validar formato de correo
    if not validar_email(email):
        flash('Formato de correo electrónico inválido', 'error')
        return redirect(url_for('index'))
    
    # Validar que la contraseña no esté vacía
    if not password:
        flash('La contraseña no puede estar vacía', 'error')
        return redirect(url_for('index'))
    
    # En un entorno real, la contraseña debería estar hasheada
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('SELECT * FROM usuarios WHERE correo = %s', (email,))
    usuario = cur.fetchone()
    cur.close()
    conn.close()
    
    if not usuario:
        flash('Usuario no encontrado', 'error')
        return redirect(url_for('index'))
    
    # En un entorno real, deberías comparar hashes, no contraseñas en texto plano
    # Aquí simulamos una verificación simple
    if usuario[5] != password:  # Suponiendo que la contraseña está en la posición 5
        flash('Contraseña incorrecta', 'error')
        return redirect(url_for('index'))
    
    # Guardar datos en la sesión
    session['user_email'] = email
    session['user_name'] = usuario[2] # Suponiendo que el nombre está en la posición 2
    
    flash('Inicio de sesión exitoso', 'success')
    return redirect(url_for('panel'))

@app.route('/contacto', methods=['POST'])
def contacto():
    nombre = request.form['nombre-contacto'].strip()
    correo = request.form['email-contacto'].strip()
    mensaje = request.form['mensaje'].strip()
    
    # Validaciones
    errores = []
    
    if not nombre:
        errores.append('El nombre es obligatorio')
    elif not validar_solo_letras(nombre):
        errores.append('El nombre solo debe contener letras y espacios')
    
    if not correo:
        errores.append('El correo electrónico es obligatorio')
    elif not validar_email(correo):
        errores.append('Formato de correo electrónico inválido')
    
    if not mensaje:
        errores.append('El mensaje es obligatorio')
    
    if errores:
        for error in errores:
            flash(error, 'error')
        return redirect(url_for('index', _anchor='contacto'))
    
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Comprobar si la tabla existe, si no, crearla
        cur.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables 
                WHERE table_name = 'contactos'
            );
        """)
        table_exists = cur.fetchone()[0]
        
        if not table_exists:
            # Crear la tabla contactos si no existe
            cur.execute("""
                CREATE TABLE contactos (
                    id SERIAL PRIMARY KEY,
                    nombre_completo VARCHAR(100) NOT NULL,
                    correo VARCHAR(100) NOT NULL,
                    mensaje TEXT NOT NULL,
                    fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            conn.commit()
        
        # Insertar el nuevo mensaje de contacto
        cur.execute('INSERT INTO contactos (nombre_completo, correo, mensaje) VALUES (%s, %s, %s)', 
                  (nombre, correo, mensaje))
        conn.commit()
        cur.close()
        conn.close()
        
        flash('Mensaje enviado correctamente. Nos pondremos en contacto pronto.', 'success')
    except Exception as e:
        flash(f'Error al enviar mensaje: {str(e)}', 'error')
    
    return redirect(url_for('index', _anchor='contacto'))

@app.route('/registro', methods=['POST'])
def registro():
    # Obtener datos del formulario
    nombres = request.form['nombre'].strip()
    apellidos = request.form['apellidos'].strip()
    correo = request.form['correo'].strip()
    telefono = request.form['telefono'].strip()
    contrasena = request.form['contrasena']
    confirmar_contrasena = request.form['confirmar_contrasena']
    
    # Validaciones
    errores = []
    
    if not nombres:
        errores.append('El nombre es obligatorio')
    elif not validar_solo_letras(nombres):
        errores.append('El nombre solo debe contener letras y espacios')
    
    if not apellidos:
        errores.append('Los apellidos son obligatorios')
    elif not validar_solo_letras(apellidos):
        errores.append('Los apellidos solo deben contener letras y espacios')
    
    if not correo:
        errores.append('El correo electrónico es obligatorio')
    elif not validar_email(correo):
        errores.append('Formato de correo electrónico inválido')
    elif existe_correo(correo):
        errores.append('Este correo electrónico ya está registrado')
    
    if not telefono:
        errores.append('El número telefónico es obligatorio')
    elif not validar_telefono(telefono):
        errores.append('El número telefónico debe tener 10 dígitos numéricos')
    
    if not contrasena:
        errores.append('La contraseña es obligatoria')
    elif not validar_contrasena(contrasena):
        errores.append('La contraseña debe tener al menos 8 caracteres y contener al menos un número')
        
    if not confirmar_contrasena:
        errores.append('Debe confirmar su contraseña')
    elif contrasena != confirmar_contrasena:
        errores.append('Las contraseñas no coinciden')
    
    if errores:
        for error in errores:
            flash(error, 'error')
        return redirect(url_for('index'))
    
    # Generar nombre de usuario a partir del correo (antes del @)
    nombre_usuario = correo.split('@')[0]
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # En un entorno real, la contraseña debería ser hasheada antes de almacenarla
        # Insertar el nuevo usuario
        cur.execute('''
            INSERT INTO usuarios 
            (correo, nombre_usuario, nombres, apellidos, numero_telefonico, contrasena)
            VALUES (%s, %s, %s, %s, %s, %s)
        ''', (correo, nombre_usuario, nombres, apellidos, telefono, contrasena))
        
        # Asignar plan básico por defecto (Guard - id_plan = 1)
        cur.execute('INSERT INTO suscripciones (correo_usuario, id_plan) VALUES (%s, %s)',
                    (correo, 1))
        
        conn.commit()
        
        # Inicio de sesión automático
        session['user_email'] = correo
        session['user_name'] = nombres
        
        flash('Registro exitoso. Bienvenido/a a DGARD.', 'success')
        return redirect(url_for('panel'))
    except psycopg2.Error as e:
        # En caso de error, hacer rollback
        conn.rollback()
        flash(f'Error en el registro: {str(e)}', 'error')
        return redirect(url_for('index'))
    finally:
        cur.close()
        conn.close()

@app.route('/perfil')
def perfil():
    if 'user_email' not in session:
        flash('Por favor inicia sesión para acceder a tu perfil', 'error')
        return redirect(url_for('index'))
    
    user_email = session['user_email']
    
    # Obtener información del usuario
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    # Consulta para información básica del usuario
    cursor.execute("SELECT * FROM usuarios WHERE correo = %s", (user_email,))
    user_info = cursor.fetchone()
    
    if not user_info:
        cursor.close()
        conn.close()
        flash('Usuario no encontrado', 'error')
        return redirect(url_for('index'))
    
    # Consulta para obtener información de la suscripción
    cursor.execute("""
        SELECT p.nombre_plan, p.numero_dispositivos, p.duracion
        FROM suscripciones s
        JOIN planes p ON s.id_plan = p.id_plan
        WHERE s.correo_usuario = %s
    """, (user_email,))
    subscription = cursor.fetchone()
    
    # Consulta para obtener las cámaras del usuario
    cursor.execute("SELECT * FROM camaras WHERE correo_usuario = %s", (user_email,))
    cameras = cursor.fetchall()
    
    # Calcular estadísticas
    total_cameras = len(cameras)
    active_cameras = total_cameras  # Suponemos que todas están activas
    
    # Obtener alertas recientes (eventos de movimiento no revisados)
    alerts_count = 0
    alerts = []
    if cameras:
        camera_ids = [camera['id_camara'] for camera in cameras]
        placeholders = ', '.join(['%s'] * len(camera_ids))
        
        # Contar alertas no revisadas
        count_query = f"""
            SELECT COUNT(*) FROM eventos_movimiento
            WHERE id_camara IN ({placeholders})
            AND revisado = FALSE
        """
        cursor.execute(count_query, camera_ids)
        alerts_count = cursor.fetchone()[0]
        
        # Obtener las últimas alertas
        query = f"""
            SELECT em.*, c.nombre_posicion
            FROM eventos_movimiento em
            JOIN camaras c ON em.id_camara = c.id_camara
            WHERE em.id_camara IN ({placeholders})
            AND em.revisado = FALSE
            ORDER BY em.fecha_evento DESC, em.hora_evento DESC
            LIMIT 5
        """
        cursor.execute(query, camera_ids)
        alerts = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    # Construir datos para enviar a la plantilla
    data = {
        'user_info': dict(user_info) if user_info else None,
        'subscription': dict(subscription) if subscription else None,
        'cameras': [dict(camera) for camera in cameras],
        'stats': {
            'total_cameras': total_cameras,
            'active_cameras': active_cameras,
            'alerts_count': alerts_count
        },
        'alerts': [dict(alert) for alert in alerts]
    }
    
    return render_template('perfil.html', data=data)

@app.route('/alertas/obtener')
def obtener_alertas():
    """API para obtener alertas no revisadas para el dropdown de notificaciones"""
    if 'user_email' not in session:
        return jsonify({'error': 'No autorizado'}), 403
    
    user_email = session['user_email']
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        
        # Obtener todas las cámaras del usuario
        cursor.execute("""
            SELECT id_camara FROM camaras 
            WHERE correo_usuario = %s
        """, (user_email,))
        user_cameras = [row['id_camara'] for row in cursor.fetchall()]
        
        if not user_cameras:
            return jsonify({'success': True, 'alerts': [], 'count': 0})
            
        placeholders = ', '.join(['%s'] * len(user_cameras))
        
        # Obtener alertas no revisadas
        cursor.execute(f"""
            SELECT em.id_evento, em.fecha_evento, em.hora_evento, em.descripcion, c.nombre_posicion
            FROM eventos_movimiento em
            JOIN camaras c ON em.id_camara = c.id_camara
            WHERE em.id_camara IN ({placeholders})
            AND em.revisado = FALSE
            ORDER BY em.fecha_evento DESC, em.hora_evento DESC
            LIMIT 5
        """, user_cameras)
        
        alerts = cursor.fetchall()
        
        # Obtener conteo total de alertas no revisadas
        cursor.execute(f"""
            SELECT COUNT(*) FROM eventos_movimiento
            WHERE id_camara IN ({placeholders})
            AND revisado = FALSE
        """, user_cameras)
        
        count = cursor.fetchone()[0]
        
        cursor.close()
        conn.close()
        
        # Formatear las alertas para JSON
        formatted_alerts = []
        for alert in alerts:
            formatted_alerts.append({
                'id_evento': alert['id_evento'],
                'nombre_posicion': alert['nombre_posicion'],
                'descripcion': alert['descripcion'],
                'fecha_evento': alert['fecha_evento'].strftime('%Y-%m-%d'),
                'hora_evento': alert['hora_evento'].strftime('%H:%M:%S')
            })
        
        return jsonify({
            'success': True,
            'alerts': formatted_alerts,
            'count': count
        })
        
    except Exception as e:
        app.logger.error(f"Error en obtener_alertas: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/alertas/marcar-revisada', methods=['POST'])
def marcar_alerta_revisada():
    """API para marcar una alerta como revisada"""
    if 'user_email' not in session:
        return jsonify({'error': 'No autorizado'}), 403
    
    user_email = session['user_email']
    
    try:
        data = request.json
        alert_id = data.get('id_evento')
        
        if not alert_id:
            return jsonify({'success': False, 'error': 'ID de alerta no proporcionado'}), 400
            
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Verificar que la alerta pertenece a una cámara del usuario
        cursor.execute("""
            SELECT em.id_evento
            FROM eventos_movimiento em
            JOIN camaras c ON em.id_camara = c.id_camara
            WHERE em.id_evento = %s
            AND c.correo_usuario = %s
        """, (alert_id, user_email))
        
        if cursor.fetchone() is None:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Alerta no encontrada o sin permisos'}), 404
        
        # Marcar la alerta como revisada
        cursor.execute("""
            UPDATE eventos_movimiento
            SET revisado = TRUE
            WHERE id_evento = %s
        """, (alert_id,))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Alerta marcada como revisada correctamente'
        })
        
    except Exception as e:
        app.logger.error(f"Error en marcar_alerta_revisada: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/configuracion', methods=['GET', 'POST'])
def configuracion():
    if 'user_email' not in session:
        flash('Por favor inicia sesión para acceder a tu configuración', 'error')
        return redirect(url_for('index'))
    
    user_email = session['user_email']
    
    if request.method == 'POST':
        # Obtenemos los datos del formulario - solo procesamos la contraseña
        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        
        # Inicializamos una lista de errores
        errores = []
        
        # Conexión a la base de datos
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        
        try:
            # Verificamos la contraseña actual
            cursor.execute("SELECT contrasena FROM usuarios WHERE correo = %s", (user_email,))
            usuario = cursor.fetchone()
            
            if not usuario:
                errores.append('Usuario no encontrado')
            elif usuario['contrasena'] != current_password:
                errores.append('La contraseña actual es incorrecta')
            
            # Validar la nueva contraseña
            if not new_password or not confirm_password:
                errores.append('Por favor complete todos los campos de contraseña')
            elif new_password != confirm_password:
                errores.append('Las nuevas contraseñas no coinciden')
            elif not validar_contrasena(new_password):
                errores.append('La nueva contraseña debe tener al menos 8 caracteres y contener al menos un número')
            
            # Si no hay errores, actualizamos la contraseña
            if not errores:
                cursor.execute("UPDATE usuarios SET contrasena = %s WHERE correo = %s", 
                               (new_password, user_email))
                conn.commit()
                flash('Contraseña actualizada correctamente', 'success')
                return redirect(url_for('perfil'))
                
        except Exception as e:
            conn.rollback()
            errores.append(f'Error: {str(e)}')
        finally:
            cursor.close()
            conn.close()
        
        # Si hay errores, los mostramos
        for error in errores:
            flash(error, 'error')
    
    # Para GET o si hay errores en POST, redirigimos al perfil
    # No es necesario una plantilla separada ya que usamos un modal en perfil.html
    return redirect(url_for('perfil'))

@app.route('/camaras')
def camaras():
    # Verificar que el usuario ha iniciado sesión
    if 'user_email' not in session:
        flash('Por favor inicia sesión para acceder a las cámaras', 'error')
        return redirect(url_for('index'))
    
    user_email = session['user_email']
    
    try:
        # Usar with para asegurar que la conexión se cierra correctamente
        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cursor:
                # Consulta para obtener todas las cámaras del usuario
                cursor.execute("""
                    SELECT id_camara, nombre_posicion, ip_camara 
                    FROM camaras 
                    WHERE correo_usuario = %s
                    ORDER BY nombre_posicion
                """, (user_email,))
                cameras = [dict(camera) for camera in cursor.fetchall()]
        
        # Construir datos para enviar a la plantilla
        data = {'cameras': cameras}
        return render_template('camaras.html', data=data)
        
    except Exception as e:
        app.logger.error(f"Error al cargar cámaras: {e}")
        flash('Error al cargar las cámaras. Por favor, inténtelo de nuevo.', 'error')
        return redirect(url_for('perfil'))

@app.route('/vigilancia')
def vigilancia():
    return redirect(url_for('camaras'))  # Redirigir a camaras por ahora

@app.route('/reportes')
def reportes():
    """Renderiza la página de reportes con datos de cámaras y estadísticas"""
    # Verificar que el usuario ha iniciado sesión
    if 'user_email' not in session:
        flash('Por favor inicia sesión para acceder a los reportes', 'error')
        return redirect(url_for('index'))
    
    user_email = session['user_email']
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        
        # Consulta para información básica del usuario
        cursor.execute("SELECT * FROM usuarios WHERE correo = %s", (user_email,))
        user_info = cursor.fetchone()
        
        # Consulta para obtener todas las cámaras del usuario
        cursor.execute("""
            SELECT id_camara, nombre_posicion, ip_camara 
            FROM camaras 
            WHERE correo_usuario = %s
            ORDER BY nombre_posicion
        """, (user_email,))
        cameras = cursor.fetchall()
        
        # Obtener estadísticas generales
        camera_ids = [camera['id_camara'] for camera in cameras]
        
        # Si no hay cámaras, establecer valores predeterminados
        if not camera_ids:
            data = {
                'user_info': dict(user_info) if user_info else {},
                'cameras': [],
                'stats': {
                    'total_alerts': 0,
                    'recent_alerts': 0,
                    'reviewed_alerts': 0,
                    'alerts_count': 0
                },
                'recent_alerts': []
            }
        else:
            # Construir placeholders para la consulta SQL
            placeholders = ', '.join(['%s'] * len(camera_ids))
            
            # Total de alertas
            cursor.execute(f"""
                SELECT COUNT(*) FROM eventos_movimiento
                WHERE id_camara IN ({placeholders})
            """, camera_ids)
            total_alerts = cursor.fetchone()[0]
            
            # Alertas recientes (últimas 24 horas)
            cursor.execute(f"""
                SELECT COUNT(*) FROM eventos_movimiento
                WHERE id_camara IN ({placeholders})
                AND fecha_evento >= CURRENT_DATE - INTERVAL '1 day'
            """, camera_ids)
            recent_alerts_count = cursor.fetchone()[0]
            
            # Alertas revisadas
            cursor.execute(f"""
                SELECT COUNT(*) FROM eventos_movimiento
                WHERE id_camara IN ({placeholders})
                AND revisado = TRUE
            """, camera_ids)
            reviewed_alerts = cursor.fetchone()[0]
            
            # Alertas no revisadas (para contador de notificaciones)
            cursor.execute(f"""
                SELECT COUNT(*) FROM eventos_movimiento
                WHERE id_camara IN ({placeholders})
                AND revisado = FALSE
            """, camera_ids)
            alerts_count = cursor.fetchone()[0]
            
            # Obtener las últimas alertas para la tabla
            cursor.execute(f"""
                SELECT em.*, c.nombre_posicion
                FROM eventos_movimiento em
                JOIN camaras c ON em.id_camara = c.id_camara
                WHERE em.id_camara IN ({placeholders})
                ORDER BY em.fecha_evento DESC, em.hora_evento DESC
                LIMIT 10
            """, camera_ids)
            recent_alerts = cursor.fetchall()
            
            # Construir el objeto de datos
            data = {
                'user_info': dict(user_info) if user_info else {},
                'cameras': [dict(camera) for camera in cameras],
                'stats': {
                    'total_alerts': total_alerts,
                    'recent_alerts': recent_alerts_count,
                    'reviewed_alerts': reviewed_alerts,
                    'alerts_count': alerts_count
                },
                'recent_alerts': [dict(alert) for alert in recent_alerts]
            }
        
        cursor.close()
        conn.close()
        
        return render_template('reportes.html', data=data)
        
    except Exception as e:
        app.logger.error(f"Error al cargar reportes: {e}")
        flash('Error al cargar los reportes. Por favor, inténtelo de nuevo.', 'error')
        return redirect(url_for('perfil'))

@app.route('/api/reportes/data')
def reportes_data():
    """API para obtener datos de reportes según filtros"""
    if 'user_email' not in session:
        return jsonify({'error': 'No autorizado'}), 403
    
    user_email = session['user_email']
    
    # Obtener parámetros de la solicitud
    date_range = request.args.get('range', 'week')
    camera_id = request.args.get('camera', 'all')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        
        # Obtener todas las cámaras del usuario
        cursor.execute("""
            SELECT id_camara FROM camaras 
            WHERE correo_usuario = %s
        """, (user_email,))
        user_cameras = [row['id_camara'] for row in cursor.fetchall()]
        
        if not user_cameras:
            return jsonify({
                'stats': {'total_alerts': 0, 'recent_alerts': 0, 'reviewed_alerts': 0},
                'charts': {
                    'byDay': {'labels': [], 'data': []},
                    'byCamera': {'labels': [], 'data': []},
                    'trend': {'labels': [], 'datasets': [{'label': 'Alertas', 'data': [], 'borderColor': 'rgba(54, 162, 235, 0.8)', 'backgroundColor': 'rgba(54, 162, 235, 0.2)'}]}
                },
                'alerts': []
            })
        
        # Filtrar por cámara específica si se proporciona
        camera_filter = []
        if camera_id != 'all' and int(camera_id) in user_cameras:
            camera_filter = [int(camera_id)]
        else:
            camera_filter = user_cameras
            
        placeholders = ', '.join(['%s'] * len(camera_filter))
        
        # Construir condición de fecha según el rango seleccionado
        date_condition = ""
        params = camera_filter.copy()  # Inicializar con las cámaras, ya que siempre serán los primeros parámetros
        
        if date_range == 'day':
            date_condition = "fecha_evento >= CURRENT_DATE - INTERVAL '1 day'"
        elif date_range == 'week':
            date_condition = "fecha_evento >= CURRENT_DATE - INTERVAL '7 days'"
        elif date_range == 'month':
            date_condition = "fecha_evento >= CURRENT_DATE - INTERVAL '30 days'"
        elif date_range == 'custom' and start_date and end_date:
            date_condition = "fecha_evento BETWEEN %s::date AND %s::date"
            params.extend([start_date, end_date])
        else:
            # Por defecto, usar una semana
            date_condition = "fecha_evento >= CURRENT_DATE - INTERVAL '7 days'"
        
        # Consulta para estadísticas
        # Total de alertas con filtros aplicados
        cursor.execute(f"""
            SELECT COUNT(*) FROM eventos_movimiento
            WHERE id_camara IN ({placeholders})
            AND {date_condition}
        """, params)
        total_alerts = cursor.fetchone()[0]
        
        # Alertas recientes (últimas 24 horas) con filtros de cámara
        cursor.execute(f"""
            SELECT COUNT(*) FROM eventos_movimiento
            WHERE id_camara IN ({placeholders})
            AND fecha_evento >= CURRENT_DATE - INTERVAL '1 day'
        """, camera_filter)
        recent_alerts = cursor.fetchone()[0]
        
        # Alertas revisadas con filtros aplicados
        cursor.execute(f"""
            SELECT COUNT(*) FROM eventos_movimiento
            WHERE id_camara IN ({placeholders})
            AND {date_condition}
            AND revisado = TRUE
        """, params)
        reviewed_alerts = cursor.fetchone()[0]
        
        # Datos para gráfica de barras (alertas por día)
        days_of_week = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
        
        cursor.execute(f"""
            SELECT 
                EXTRACT(DOW FROM fecha_evento) as day_number,
                COUNT(*) as count
            FROM eventos_movimiento
            WHERE id_camara IN ({placeholders})
            AND {date_condition}
            GROUP BY day_number
            ORDER BY day_number
        """, params)
        
        by_day_data = cursor.fetchall()
        
        # Inicializar contadores para cada día
        daily_counts = [0] * 7
        
        # PostgreSQL EXTRACT(DOW) devuelve 0 para domingo y 6 para sábado
        # Ajustar para que domingo sea 6 (último día) y lunes sea 0 (primer día)
        for row in by_day_data:
            day_num = int(row['day_number'])
            # Convertir de domingo=0 a domingo=6
            adjusted_day = (day_num - 1) % 7
            daily_counts[adjusted_day] = row['count']
        
        # Datos para gráfica circular (distribución por cámara)
        cursor.execute(f"""
            SELECT 
                c.nombre_posicion,
                COUNT(e.id_evento) as total
            FROM camaras c
            LEFT JOIN eventos_movimiento e ON c.id_camara = e.id_camara
            AND {date_condition}
            WHERE c.id_camara IN ({placeholders})
            GROUP BY c.nombre_posicion
            ORDER BY total DESC
        """, params)
        
        camera_distribution = cursor.fetchall()
        camera_labels = [row['nombre_posicion'] for row in camera_distribution]
        camera_data = [row['total'] for row in camera_distribution]
        
        # Datos para gráfica de tendencia (últimas 4 semanas)
        cursor.execute(f"""
            SELECT 
                EXTRACT(WEEK FROM fecha_evento) as week_number,
                COUNT(*) as count
            FROM eventos_movimiento
            WHERE id_camara IN ({placeholders})
            AND fecha_evento >= CURRENT_DATE - INTERVAL '28 days'
            GROUP BY week_number
            ORDER BY week_number
        """, camera_filter)
        
        trend_data = cursor.fetchall()
        
        # Preparar etiquetas para semanas (últimas 4 semanas)
        current_week = datetime.now().isocalendar()[1]
        week_labels = []
        week_counts = []
        
        for i in range(4):
            week_num = (current_week - 3 + i) % 52
            if week_num <= 0:
                week_num += 52
            week_labels.append(f'Semana {week_num}')
            
            # Buscar si hay datos para esta semana
            count = 0
            for row in trend_data:
                if int(row['week_number']) == week_num:
                    count = row['count']
                    break
            week_counts.append(count)
        
        # Últimas alertas para la tabla
        cursor.execute(f"""
            SELECT em.*, c.nombre_posicion
            FROM eventos_movimiento em
            JOIN camaras c ON em.id_camara = c.id_camara
            WHERE em.id_camara IN ({placeholders})
            AND {date_condition}
            ORDER BY em.fecha_evento DESC, em.hora_evento DESC
            LIMIT 10
        """, params)
        
        alerts = cursor.fetchall()
        
        # Formatear las alertas para JSON
        formatted_alerts = []
        for alert in alerts:
            formatted_alerts.append({
                'nombre_posicion': alert['nombre_posicion'],
                'descripcion': alert['descripcion'],
                'fecha_evento': alert['fecha_evento'].strftime('%Y-%m-%d'),
                'hora_evento': alert['hora_evento'].strftime('%H:%M:%S'),
                'revisado': alert['revisado']
            })
        
        cursor.close()
        conn.close()
        
        # Construir respuesta JSON
        return jsonify({
            'stats': {
                'total_alerts': total_alerts,
                'recent_alerts': recent_alerts,
                'reviewed_alerts': reviewed_alerts
            },
            'charts': {
                'byDay': {
                    'labels': days_of_week,
                    'data': daily_counts
                },
                'byCamera': {
                    'labels': camera_labels,
                    'data': camera_data
                },
                'trend': {
                    'labels': week_labels,
                    'datasets': [{
                        'label': 'Alertas',
                        'data': week_counts,
                        'borderColor': 'rgba(54, 162, 235, 0.8)',
                        'backgroundColor': 'rgba(54, 162, 235, 0.2)'
                    }]
                }
            },
            'alerts': formatted_alerts
        })
        
    except Exception as e:
        app.logger.error(f"Error en reportes_data: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/reportes/duracion')
def reportes_duracion():
    """API para obtener datos de duración de movimiento según filtros"""
    if 'user_email' not in session:
        return jsonify({'error': 'No autorizado'}), 403
    
    user_email = session['user_email']
    
    # Obtener parámetros de la solicitud
    date_range = request.args.get('range', 'week')
    camera_id = request.args.get('camera', 'all')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        
        # Obtener todas las cámaras del usuario
        cursor.execute("""
            SELECT id_camara FROM camaras 
            WHERE correo_usuario = %s
        """, (user_email,))
        user_cameras = [row['id_camara'] for row in cursor.fetchall()]
        
        if not user_cameras:
            return jsonify({
                'stats': {'bajo': 0, 'moderado': 0, 'alto': 0, 'critico': 0},
                'promedios': [],
                'eventos': []
            })
        
        # Filtrar por cámara específica si se proporciona
        camera_filter = []
        if camera_id != 'all' and int(camera_id) in user_cameras:
            camera_filter = [int(camera_id)]
        else:
            camera_filter = user_cameras
            
        placeholders = ', '.join(['%s'] * len(camera_filter))
        
        # Construir condición de fecha según el rango seleccionado
        date_condition = ""
        params = camera_filter.copy()  # Inicializar con las cámaras, ya que siempre serán los primeros parámetros
        
        if date_range == 'day':
            date_condition = "em.fecha_evento >= CURRENT_DATE - INTERVAL '1 day'"
        elif date_range == 'week':
            date_condition = "em.fecha_evento >= CURRENT_DATE - INTERVAL '7 days'"
        elif date_range == 'month':
            date_condition = "em.fecha_evento >= CURRENT_DATE - INTERVAL '30 days'"
        elif date_range == 'custom' and start_date and end_date:
            date_condition = "em.fecha_evento BETWEEN %s::date AND %s::date"
            params.extend([start_date, end_date])
        else:
            # Por defecto, usar una semana
            date_condition = "em.fecha_evento >= CURRENT_DATE - INTERVAL '7 days'"
        
        # 1. Obtener estadísticas de categorización por duración
        cursor.execute(f"""
            SELECT 
                COUNT(CASE WHEN dm.duracion_segundos < 5 THEN 1 END) as bajo,
                COUNT(CASE WHEN dm.duracion_segundos >= 5 AND dm.duracion_segundos < 20 THEN 1 END) as moderado,
                COUNT(CASE WHEN dm.duracion_segundos >= 20 AND dm.duracion_segundos < 60 THEN 1 END) as alto,
                COUNT(CASE WHEN dm.duracion_segundos >= 60 THEN 1 END) as critico
            FROM duracion_movimiento dm
            JOIN eventos_movimiento em ON dm.id_evento = em.id_evento
            WHERE em.id_camara IN ({placeholders})
            AND {date_condition}
        """, params)
        
        stats = cursor.fetchone()
        
        # 2. Obtener duración promedio por cámara
        cursor.execute(f"""
            SELECT 
                c.nombre_posicion,
                COALESCE(AVG(dm.duracion_segundos), 0) as duracion_promedio
            FROM camaras c
            LEFT JOIN eventos_movimiento em ON c.id_camara = em.id_camara
            LEFT JOIN duracion_movimiento dm ON em.id_evento = dm.id_evento
            WHERE c.id_camara IN ({placeholders})
            AND ({date_condition} OR em.id_evento IS NULL)
            GROUP BY c.nombre_posicion
            ORDER BY duracion_promedio DESC
        """, params)
        
        promedios = cursor.fetchall()
        
        # 3. Obtener eventos de larga duración (> 20 segundos)
        cursor.execute(f"""
            SELECT 
                c.nombre_posicion,
                em.fecha_evento,
                em.hora_evento,
                dm.duracion_segundos,
                em.revisado
            FROM duracion_movimiento dm
            JOIN eventos_movimiento em ON dm.id_evento = em.id_evento
            JOIN camaras c ON em.id_camara = c.id_camara
            WHERE em.id_camara IN ({placeholders})
            AND {date_condition}
            AND dm.duracion_segundos >= 20
            ORDER BY dm.duracion_segundos DESC, em.fecha_evento DESC, em.hora_evento DESC
            LIMIT 10
        """, params)
        
        eventos = cursor.fetchall()
        
        # Formatear los datos para JSON
        formatted_stats = {
            'bajo': stats['bajo'] if stats else 0,
            'moderado': stats['moderado'] if stats else 0,
            'alto': stats['alto'] if stats else 0,
            'critico': stats['critico'] if stats else 0
        }
        
        formatted_promedios = []
        for promedio in promedios:
            formatted_promedios.append({
                'nombre_posicion': promedio['nombre_posicion'],
                'duracion_promedio': round(float(promedio['duracion_promedio']), 1)
            })
        
        formatted_eventos = []
        for evento in eventos:
            formatted_eventos.append({
                'nombre_posicion': evento['nombre_posicion'],
                'fecha_evento': evento['fecha_evento'].strftime('%Y-%m-%d'),
                'hora_evento': evento['hora_evento'].strftime('%H:%M:%S'),
                'duracion_segundos': evento['duracion_segundos'],
                'revisado': evento['revisado']
            })
        
        cursor.close()
        conn.close()
        
        # Construir respuesta JSON
        return jsonify({
            'stats': formatted_stats,
            'promedios': formatted_promedios,
            'eventos': formatted_eventos
        })
        
    except Exception as e:
        app.logger.error(f"Error en reportes_duracion: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/reportes/generar', methods=['POST'])
def generar_reporte():
    if 'user_email' not in session:
        return jsonify({'error': 'No autorizado'}), 403

    user_email = session['user_email']
    data = request.get_json()

    if not data:
        return jsonify({'error': 'Datos no válidos'}), 400

    title = data.get('title', 'Reporte de alertas')
    report_type = data.get('type', 'complete')
    format_type = data.get('format', 'pdf')
    include_graphs = data.get('includeGraphs', True)
    include_alerts = data.get('includeAlerts', True)
    include_summary = data.get('includeSummary', True)
    date_range = data.get('dateRange', 'week')
    camera_id = data.get('cameraId', 'all')
    start_date = data.get('startDate')
    end_date = data.get('endDate')

    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

        cursor.execute("""
            SELECT nombre_usuario, nombres, apellidos
            FROM usuarios WHERE correo = %s
        """, (user_email,))
        user_info = cursor.fetchone()

        if not user_info:
            return jsonify({'error': 'Usuario no encontrado'}), 404

        cursor.execute("""
            SELECT id_camara, nombre_posicion FROM camaras 
            WHERE correo_usuario = %s
        """, (user_email,))
        cameras = cursor.fetchall()

        camera_filter = []
        if camera_id != 'all' and any(c['id_camara'] == int(camera_id) for c in cameras):
            camera_filter = [int(camera_id)]
        else:
            camera_filter = [c['id_camara'] for c in cameras]

        if not camera_filter:
            return jsonify({'error': 'No hay cámaras disponibles'}), 404

        placeholders = ', '.join(['%s'] * len(camera_filter))

        date_condition = ""
        params = []

        if date_range == 'day':
            date_condition = "fecha_evento >= CURRENT_DATE - INTERVAL '1 day'"
        elif date_range == 'week':
            date_condition = "fecha_evento >= CURRENT_DATE - INTERVAL '7 days'"
        elif date_range == 'month':
            date_condition = "fecha_evento >= CURRENT_DATE - INTERVAL '30 days'"
        elif date_range == 'custom' and start_date and end_date:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
            date_condition = "fecha_evento BETWEEN %s AND %s"
            params = [start_date_obj, end_date_obj]
        else:
            date_condition = "fecha_evento >= CURRENT_DATE - INTERVAL '7 days'"

        full_params = camera_filter.copy()
        if params:
            full_params.extend(params)

        stats = {}
        if include_summary:
            cursor.execute(f"""
                SELECT COUNT(*) FROM eventos_movimiento
                WHERE id_camara IN ({placeholders})
                AND {date_condition}
            """, full_params)
            stats['total_alerts'] = cursor.fetchone()[0]

            cursor.execute(f"""
                SELECT COUNT(*) FROM eventos_movimiento
                WHERE id_camara IN ({placeholders})
                AND fecha_evento >= CURRENT_DATE - INTERVAL '1 day'
            """, camera_filter)
            stats['recent_alerts'] = cursor.fetchone()[0]

            cursor.execute(f"""
                SELECT COUNT(*) FROM eventos_movimiento
                WHERE id_camara IN ({placeholders})
                AND {date_condition}
                AND revisado = TRUE
            """, full_params)
            stats['reviewed_alerts'] = cursor.fetchone()[0]

            # Estadísticas por duración
            cursor.execute(f"""
                SELECT 
                    COUNT(CASE WHEN dm.duracion_segundos < 5 THEN 1 END) as bajo,
                    COUNT(CASE WHEN dm.duracion_segundos >= 5 AND dm.duracion_segundos < 20 THEN 1 END) as moderado,
                    COUNT(CASE WHEN dm.duracion_segundos >= 20 AND dm.duracion_segundos < 60 THEN 1 END) as alto,
                    COUNT(CASE WHEN dm.duracion_segundos >= 60 THEN 1 END) as critico
                FROM duracion_movimiento dm
                JOIN eventos_movimiento em ON dm.id_evento = em.id_evento
                WHERE em.id_camara IN ({placeholders})
                AND {date_condition}
            """, full_params)

            duracion_stats = cursor.fetchone()
            stats['duracion_categorias'] = {
                'bajo': duracion_stats['bajo'] if duracion_stats else 0,
                'moderado': duracion_stats['moderado'] if duracion_stats else 0,
                'alto': duracion_stats['alto'] if duracion_stats else 0,
                'critico': duracion_stats['critico'] if duracion_stats else 0
            }

            cursor.execute(f"""
                SELECT 
                    c.nombre_posicion,
                    COALESCE(AVG(dm.duracion_segundos), 0) as duracion_promedio
                FROM camaras c
                LEFT JOIN eventos_movimiento em ON c.id_camara = em.id_camara
                LEFT JOIN duracion_movimiento dm ON em.id_evento = dm.id_evento
                WHERE c.id_camara IN ({placeholders})
                AND ({date_condition} OR em.id_evento IS NULL)
                GROUP BY c.nombre_posicion
                ORDER BY duracion_promedio DESC
            """, full_params)

            stats['duracion_promedio'] = [
                {
                    'nombre_posicion': row['nombre_posicion'],
                    'duracion_promedio': round(float(row['duracion_promedio']), 1)
                }
                for row in cursor.fetchall()
            ]

            cursor.execute(f"""
                SELECT 
                    c.nombre_posicion,
                    em.fecha_evento,
                    em.hora_evento,
                    dm.duracion_segundos,
                    em.revisado
                FROM duracion_movimiento dm
                JOIN eventos_movimiento em ON dm.id_evento = em.id_evento
                JOIN camaras c ON em.id_camara = c.id_camara
                WHERE em.id_camara IN ({placeholders})
                AND {date_condition}
                AND dm.duracion_segundos >= 20
                ORDER BY dm.duracion_segundos DESC, em.fecha_evento DESC, em.hora_evento DESC
                LIMIT 10
            """, full_params)

            stats['eventos_largos'] = [
                {
                    'nombre_posicion': row['nombre_posicion'],
                    'fecha_evento': row['fecha_evento'].strftime('%Y-%m-%d'),
                    'hora_evento': row['hora_evento'].strftime('%H:%M:%S'),
                    'duracion_segundos': row['duracion_segundos'],
                    'revisado': row['revisado']
                }
                for row in cursor.fetchall()
            ]

        charts_data = {}
        if include_graphs:
            days_of_week = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
            cursor.execute(f"""
                SELECT 
                    EXTRACT(DOW FROM fecha_evento) as day_number,
                    COUNT(*) as count
                FROM eventos_movimiento
                WHERE id_camara IN ({placeholders})
                AND {date_condition}
                GROUP BY day_number
                ORDER BY day_number
            """, full_params)

            by_day_data = cursor.fetchall()
            daily_counts = [0] * 7
            for row in by_day_data:
                day_num = int(row['day_number'])
                adjusted_day = (day_num - 1) % 7
                daily_counts[adjusted_day] = row['count']

            charts_data['by_day'] = {
                'labels': days_of_week,
                'data': daily_counts
            }

            cursor.execute(f"""
                SELECT 
                    c.nombre_posicion,
                    COUNT(e.id_evento) as total
                FROM camaras c
                LEFT JOIN eventos_movimiento e ON c.id_camara = e.id_camara
                AND {date_condition}
                WHERE c.id_camara IN ({placeholders})
                GROUP BY c.nombre_posicion
                ORDER BY total DESC
            """, full_params)

            cam_data = cursor.fetchall()
            charts_data['by_camera'] = {
                'labels': [row['nombre_posicion'] for row in cam_data],
                'data': [row['total'] for row in cam_data]
            }

        alerts_data = []
        if include_alerts:
            limit_clause = "LIMIT 50" if report_type == 'alerts-summary' else "LIMIT 20"
            cursor.execute(f"""
                SELECT em.*, c.nombre_posicion
                FROM eventos_movimiento em
                JOIN camaras c ON em.id_camara = c.id_camara
                WHERE em.id_camara IN ({placeholders})
                AND {date_condition}
                ORDER BY em.fecha_evento DESC, em.hora_evento DESC
                {limit_clause}
            """, full_params)

            alerts_data = [
                {
                    'id_evento': row['id_evento'],
                    'nombre_posicion': row['nombre_posicion'],
                    'descripcion': row['descripcion'],
                    'fecha_evento': row['fecha_evento'].strftime('%Y-%m-%d'),
                    'hora_evento': row['hora_evento'].strftime('%H:%M:%S'),
                    'revisado': row['revisado']
                }
                for row in cursor.fetchall()
            ]
        cursor.close()
        conn.close()

        safe_title = re.sub(r'[^\w\s-]', '', title).strip().replace(' ', '_')
        filename = f"reporte_{safe_title}_{user_info['nombre_usuario']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        file_path = ""

        if format_type == 'pdf':
            file_path = generate_pdf_report(title, user_info, stats, charts_data, alerts_data, report_type, date_range, start_date, end_date, filename)
        elif format_type == 'excel':
            file_path = generate_excel_report(title, user_info, stats, charts_data, alerts_data, report_type, date_range, start_date, end_date, filename)
        elif format_type == 'csv':
            file_path = generate_csv_report(title, user_info, stats, charts_data, alerts_data, report_type, date_range, start_date, end_date, filename)

        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            camera_ref = camera_filter[0] if camera_filter else None
            cursor.execute("""
                INSERT INTO reportes (id_camara, fecha_reporte, hora_reporte, descripcion)
                VALUES (%s, CURRENT_DATE, CURRENT_TIME, %s)
            """, (camera_ref, f"Reporte generado: {title}"))
            conn.commit()
            cursor.close()
            conn.close()
        except Exception as e:
            app.logger.error(f"Error al registrar reporte: {str(e)}")

        file_url = url_for('download_report', filename=os.path.basename(file_path), _external=True)
        return jsonify({
            'success': True,
            'fileUrl': file_url,
            'message': f'Reporte {title} generado exitosamente'
        })

    except Exception as e:
        app.logger.error(f"Error en generar_reporte: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/reports/download/<filename>')
def download_report(filename):
    reports_dir = os.path.join(app.static_folder, 'reports')
    return send_from_directory(reports_dir, filename, as_attachment=True)

@app.route('/logout')
def logout():
    # Eliminar el email de la sesión
    session.pop('user_email', None)
    flash('Has cerrado sesión correctamente', 'success')
    return redirect(url_for('index'))


def get_db_connection():
    
    conn = psycopg2.connect(
        host="localhost",
        database="DGard",
        user="postgres",
        password=""
    )
    return conn

def validar_email(email):
    """Validar formato de correo electrónico"""
    patron = r'^[a-zA-Z0-9._-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,6}$'
    return re.match(patron, email) is not None

def validar_solo_letras(texto):
    """Validar que el texto solo contenga letras y espacios"""
    patron = r'^[a-zA-ZáéíóúÁÉÍÓÚñÑüÜ\s]+$'
    return re.match(patron, texto) is not None

def validar_telefono(telefono):
    """Validar que el teléfono tenga 10 dígitos numéricos (formato mexicano)"""
    patron = r'^[0-9]{10}$'
    return re.match(patron, telefono) is not None

def validar_contrasena(contrasena):
    """Validar que la contraseña tenga al menos 8 caracteres y un número"""
    if len(contrasena) < 8:
        return False
    return any(char.isdigit() for char in contrasena)

def existe_correo(correo):
    """Verificar si el correo ya existe en la base de datos"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('SELECT COUNT(*) FROM usuarios WHERE correo = %s', (correo,))
    count = cur.fetchone()[0]
    cur.close()
    conn.close()
    return count > 0

def get_unique_filename(base_path, file_extension):
    """Genera un nombre de archivo único si ya existe uno con el mismo nombre"""
    counter = 1
    file_path = f"{base_path}.{file_extension}"
    
    while os.path.exists(file_path):
        file_path = f"{base_path}_{counter}.{file_extension}"
        counter += 1
        
    return file_path

def generate_pdf_report(title, user_info, stats, charts_data, alerts_data, report_type, date_range, start_date, end_date, filename):
    import os
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import numpy as np
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.platypus import PageBreak
    import io
    from datetime import datetime
    
    # Crear directorio para reportes si no existe
    reports_dir = os.path.join(app.static_folder, 'reports')
    if not os.path.exists(reports_dir):
        os.makedirs(reports_dir)
    
    # Obtener una ruta de archivo única
    base_path = os.path.join(reports_dir, filename)
    file_path = get_unique_filename(base_path, 'pdf')
    
    # Determinar la orientación según el tipo de reporte
    pagesize = letter
    if report_type == "cameras-analysis" or report_type == "complete-integrated":
        pagesize = landscape(letter)
    
    # Crear el documento
    doc = SimpleDocTemplate(file_path, pagesize=pagesize, leftMargin=0.5*inch, rightMargin=0.5*inch)
    styles = getSampleStyleSheet()
    elements = []
    
    # ================== MEJORA 1: ESTILOS PERSONALIZADOS MÁS MODERNOS Y ATRACTIVOS ==================
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        alignment=TA_CENTER,
        spaceAfter=16,
        textColor=colors.HexColor('#1A365D'),  # Azul más oscuro y profesional
        leading=24,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=16,
        spaceBefore=12,
        spaceAfter=12,
        textColor=colors.HexColor('#2A4365'),  # Azul complementario
        leading=20,
        fontName='Helvetica-Bold'
    )
    
    section_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading3'],
        fontSize=14,
        spaceBefore=10,
        spaceAfter=10,
        textColor=colors.HexColor('#2C5282'),  # Azul medio
        leading=16,
        fontName='Helvetica-Bold'
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=11,
        spaceAfter=8,
        leading=14,
        fontName='Helvetica'
    )
    
    # ================== MEJORA 2: CABECERA DEL REPORTE MÁS PROFESIONAL ==================
    # Título y fecha en un diseño atractivo usando una tabla
    # Crear marco superior con línea de color
    header_line = Table([[""]],
                       colWidths=[doc.width],
                       rowHeights=[0.1*inch])
    header_line.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1A365D')),
        ('LINEABOVE', (0, 0), (-1, -1), 2, colors.HexColor('#1A365D')),
    ]))
    elements.append(header_line)
    elements.append(Spacer(1, 0.2*inch))
    
    # Título y subtítulo según el tipo de reporte
    if report_type == "cameras-analysis":
        main_title = f"{title} - ANÁLISIS DE CÁMARAS"
        icon_text = "📹"  # Icono relacionado con cámaras
    elif report_type == "sensors-duration":
        main_title = f"{title} - ANÁLISIS DE SENSORES Y DURACIÓN"
        icon_text = "⏱️"  # Icono relacionado con duración/tiempo
    elif report_type == "complete-integrated":
        main_title = f"{title} - REPORTE COMPLETO INTEGRADO"
        icon_text = "📊"  # Icono relacionado con reporte completo
    else:
        main_title = title
        icon_text = "📄"  # Icono genérico para documento
    
    # Crear encabezado con más estilo
    header_data = [[f"{icon_text} {main_title}"]]
    header_table = Table(header_data, colWidths=[doc.width])
    header_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1A365D')),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 18),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(header_table)
    
    # Fecha y hora de generación con mejor formato
    current_time = datetime.now().strftime('%d/%m/%Y - %H:%M:%S')
    date_data = [[f"Generado el: {current_time}"]]
    date_table = Table(date_data, colWidths=[doc.width])
    date_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#4A5568')),  # Gris oscuro
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Oblique'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
    ]))
    elements.append(date_table)
    
    # Línea separadora
    separator_line = Table([[""]],
                          colWidths=[doc.width],
                          rowHeights=[0.5])
    separator_line.setStyle(TableStyle([
        ('LINEBELOW', (0, 0), (-1, 0), 0.5, colors.HexColor('#CBD5E0')),  # Gris claro
    ]))
    elements.append(separator_line)
    elements.append(Spacer(1, 0.3*inch))
    
    # ================== MEJORA 3: INFORMACIÓN DEL USUARIO EN FORMATO MÁS MODERNO ==================
    elements.append(Paragraph("Información del Usuario", section_style))
    
    # Creamos una tabla más moderna para la información del usuario
    user_data = [
        ["Nombre:", f"{user_info['nombres']} {user_info['apellidos']}"],
        ["Usuario:", user_info['nombre_usuario']]
    ]
    user_table = Table(user_data, colWidths=[1.5*inch, 4.5*inch])
    user_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#EBF4FF')),  # Azul muy claro
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#2C5282')),  # Azul medio
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (0, -1), 12),  # Padding izquierdo para las etiquetas
        ('RIGHTPADDING', (0, 0), (0, -1), 12),  # Padding derecho para las etiquetas
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E0'))
    ]))
    elements.append(user_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # ================== MEJORA 4: PERÍODO DEL REPORTE MÁS VISUAL ==================
    # Crear una visualización más atractiva del período analizado
    period_text = "Período analizado: "
    period_icon = ""
    
    if date_range == 'day':
        period_text += "Último día"
        period_icon = "📅"
    elif date_range == 'week':
        period_text += "Última semana"
        period_icon = "📆"
    elif date_range == 'month':
        period_text += "Último mes"
        period_icon = "📅"
    elif date_range == 'custom' and start_date and end_date:
        period_text += f"Del {start_date} al {end_date}"
        period_icon = "🗓️"
    
    period_data = [[f"{period_icon} {period_text}"]]
    period_table = Table(period_data, colWidths=[doc.width])
    period_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#EBF8FF')),  # Azul muy claro
        ('ALIGN', (0, 0), (0, 0), 'CENTER'),
        ('TEXTCOLOR', (0, 0), (0, 0), colors.HexColor('#2B6CB0')),  # Azul
        ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (0, 0), 12),
        ('BOTTOMPADDING', (0, 0), (0, 0), 8),
        ('TOPPADDING', (0, 0), (0, 0), 8),
        ('ROUND', (0, 0), (0, 0), 6),  # Bordes redondeados
        ('BOX', (0, 0), (0, 0), 0.5, colors.HexColor('#90CDF4'))  # Borde azul claro
    ]))
    elements.append(period_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # ================== REPORTE ESPECÍFICO: ANÁLISIS DE SENSORES Y DURACIÓN ==================
    if report_type == "sensors-duration":
        # 1.1 Sección de Resumen de Duración
        if stats and 'duracion_categorias' in stats:
            # Título de la sección con estilo mejorado y separador
            section_title_table = Table([["Análisis de Duración de Eventos"]],
                                        colWidths=[doc.width])
            section_title_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#2B6CB0')),  # Azul
                ('TEXTCOLOR', (0, 0), (0, 0), colors.white),
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (0, 0), 14),
                ('BOTTOMPADDING', (0, 0), (0, 0), 8),
                ('TOPPADDING', (0, 0), (0, 0), 8),
                ('LEFTPADDING', (0, 0), (0, 0), 12),
                ('ROUND', (0, 0), (0, 0), 4),  # Bordes redondeados
            ]))
            elements.append(section_title_table)
            elements.append(Spacer(1, 0.2*inch))
            
            # Texto explicativo mejorado con iconos visuales
            duration_explanation = """
            <para>El análisis de duración clasifica los eventos en categorías según el tiempo que permanecen activos:
            <br/><b>• <font color="#66BB6A">Bajo:</font></b> Menos de 5 segundos - Eventos breves, generalmente falsos positivos.
            <br/><b>• <font color="#FFC107">Moderado:</font></b> Entre 5 y 20 segundos - Duración normal para tránsito regular.
            <br/><b>• <font color="#FF9800">Alto:</font></b> Entre 20 y 60 segundos - Actividad prolongada que merece revisión.
            <br/><b>• <font color="#F44336">Crítico:</font></b> Más de 60 segundos - Posible situación de riesgo que requiere atención inmediata.
            </para>
            """
            elements.append(Paragraph(duration_explanation, normal_style))
            elements.append(Spacer(1, 0.2*inch))
            
            # Gráfico de distribución por categorías de duración MEJORADO
            plt.figure(figsize=(7, 4.5))
            categories = ['Bajo (<5s)', 'Moderado (5-20s)', 'Alto (20-60s)', 'Crítico (>60s)']
            values = [
                stats['duracion_categorias']['bajo'],
                stats['duracion_categorias']['moderado'],
                stats['duracion_categorias']['alto'],
                stats['duracion_categorias']['critico']
            ]
            
            # Colores para indicar severidad con un toque más vibrante pero profesional
            bar_colors = ['#66BB6A', '#FFC107', '#FF9800', '#F44336']
            
            # Crear gráfico de barras con estilo mejorado
            ax = plt.subplot(111)
            bars = plt.bar(categories, values, color=bar_colors, width=0.65)
            
            # Añadir valores sobre las barras
            for bar in bars:
                height = bar.get_height()
                plt.text(bar.get_x() + bar.get_width()/2., height + 0.8,
                       f'{int(height)}', ha='center', va='bottom', fontweight='bold', fontsize=12)
            
            # Mejorar diseño del gráfico
            plt.title('Distribución de Eventos por Duración', fontsize=16, pad=20, fontweight='bold')
            plt.ylabel('Número de eventos', fontsize=12)
            plt.grid(axis='y', linestyle='--', alpha=0.3)
            
            # Eliminar bordes innecesarios
            for spine in ax.spines.values():
                spine.set_visible(False)
            
            # Añadir colores de fondo para mejor visualización
            ax.set_facecolor('#F7FAFC')  # Fondo gris muy claro
            plt.gcf().set_facecolor('#FFFFFF')
            
            # Ajustar ticks del eje Y para mejor legibilidad
            plt.yticks(fontsize=8)
            plt.xticks(fontsize=9, fontweight='bold')
            
            # Añadir valor total como anotación
            total = sum(values)
            plt.annotate(f'Total: {total} eventos', 
                       xy=(0.5, 0.97),
                       xycoords='axes fraction',
                       fontsize=12,
                       fontweight='bold',
                       ha='center',
                       va='top',
                       bbox=dict(boxstyle="round,pad=0.3", fc="#EBF8FF", ec="#4299E1", alpha=0.8))
            
            plt.tight_layout()
            
            # Guardar gráfica
            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
            img_buffer.seek(0)
            
            # Crear imagen para el PDF
            img = Image(img_buffer)
            img.drawHeight = 3.5*inch
            img.drawWidth = 7*inch
            elements.append(img)
            plt.close('all')
            
            elements.append(Spacer(1, 0.3*inch))
            
            # Tabla resumen con porcentajes - MEJORADA
            total_events = sum(values)
            duration_data = [
                ["Categoría", "Eventos", "Porcentaje", "Nivel de Atención"]
            ]
            
            # Definiciones de niveles de atención para cada categoría
            attention_levels = [
                "Bajo - Revisión opcional",
                "Moderado - Revisión recomendada",
                "Alto - Requiere revisión",
                "Crítico - Atención inmediata"
            ]
            
            for i, cat in enumerate(categories):
                count = values[i]
                percentage = (count/total_events*100) if total_events > 0 else 0
                duration_data.append([
                    cat,
                    str(count),
                    f"{percentage:.1f}%",
                    attention_levels[i]
                ])
            
            # Añadir fila de totales
            duration_data.append([
                "TOTAL",
                str(total_events),
                "100.0%",
                ""
            ])
            
            # Crear tabla con mejor diseño
            duration_table = Table(duration_data, colWidths=[1.5*inch, 1*inch, 1*inch, 3.5*inch])
            
            # Estilo de la tabla más moderno
            duration_style = [
                # Cabecera
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A365D')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (2, -1), 'CENTER'),
                ('ALIGN', (3, 0), (3, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E0')),
                # Configuraciones generales
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 1), (-1, -2), 8),
                ('TOPPADDING', (0, 1), (-1, -2), 8),
            ]
            
            # Mejorar colores para categorías - contraste mejorado
            color_hex = ['#E8F5E9', '#FFF8E1', '#FFF3E0', '#FFEBEE']  # Fondos claros para mejor legibilidad
            text_colors = ['#2E7D32', '#F57F17', '#E65100', '#B71C1C']  # Textos con contraste
            
            for i in range(4):
                row = i + 1
                duration_style.append(('BACKGROUND', (0, row), (0, row), colors.HexColor(color_hex[i])))
                duration_style.append(('TEXTCOLOR', (0, row), (0, row), colors.HexColor(text_colors[i])))
                duration_style.append(('FONTNAME', (0, row), (0, row), 'Helvetica-Bold'))
            
            # Fila de totales con estilo distintivo
            duration_style.append(('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#EDF2F7')))
            duration_style.append(('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'))
            duration_style.append(('LINEBELOW', (0, -2), (-1, -2), 1, colors.HexColor('#4A5568')))
            duration_style.append(('TOPPADDING', (0, -1), (-1, -1), 10))
            duration_style.append(('BOTTOMPADDING', (0, -1), (-1, -1), 10))
            
            duration_table.setStyle(TableStyle(duration_style))
            elements.append(duration_table)
            elements.append(Spacer(1, 0.5*inch))
        
           # 1.2 Duración promedio por ubicación
        if stats and 'duracion_promedio' in stats and len(stats['duracion_promedio']) > 0:
            elements.append(Paragraph("Duración Promedio por Ubicación", subtitle_style))

            avg_duration_text = """
            <para>El análisis de duración promedio por ubicación permite identificar patrones espaciales en la activación
            de sensores. Las ubicaciones con valores más altos pueden requerir revisión de configuración o indicar
            situaciones que demandan atención prioritaria.</para>
            """
            elements.append(Paragraph(avg_duration_text, normal_style))
            elements.append(Spacer(1, 0.2 * inch))

            locations = [item['nombre_posicion'] for item in stats['duracion_promedio']]
            avg_durations = [item['duracion_promedio'] for item in stats['duracion_promedio']]

        if len(locations) > 10:
            sorted_data = sorted(zip(locations, avg_durations), key=lambda x: x[1], reverse=True)
            locations = [item[0] for item in sorted_data[:10]]
            avg_durations = [item[1] for item in sorted_data[:10]]
            note = "Nota: Se muestran las 10 ubicaciones con mayor duración promedio."
            elements.append(Paragraph(note, ParagraphStyle(
            'Note',
            parent=normal_style,
            fontSize=9,
            textColor=colors.HexColor('#718096'),
            fontName='Helvetica-Oblique'
            )))

            plt.figure(figsize=(5.5, max(6, len(locations) * 0.9)), dpi=150)
            plt.style.use('seaborn-v0_8-whitegrid')

        bar_colors = []
        for duration in avg_durations:
            if duration < 5:
             bar_colors.append('#38A169')
            elif duration < 20:
             bar_colors.append('#DD6B20')
            elif duration < 60:
             bar_colors.append('#E53E3E')
            else:
             bar_colors.append('#822727')

            bars = plt.barh(locations, avg_durations, color=bar_colors, height=0.7,
                edgecolor='white', linewidth=0.7)

        for bar in bars:
            width = bar.get_width()
        plt.text(width + max(avg_durations) * 0.02, bar.get_y() + bar.get_height() / 2,
                 f'{width:.1f}s', ha='left', va='center',
                 fontsize=9, fontweight='bold', color='#2D3748')

        plt.title('Duración Promedio de Eventos por Ubicación',
              fontsize=14, fontweight='bold', pad=15, color='#2D3748')
        plt.xlabel('Duración (segundos)', fontsize=10, color='#4A5568', labelpad=10)

        plt.gca().spines['top'].set_visible(False)
        plt.gca().spines['right'].set_visible(False)
        plt.gca().spines['left'].set_color('#CBD5E0')
        plt.gca().spines['bottom'].set_color('#CBD5E0')

        plt.grid(axis='x', linestyle='--', alpha=0.6, color='#E2E8F0')

        max_x = max(avg_durations) * 1.2
        plt.xlim(0, max_x)

        x_ticks = [0, 5, 20, 60]
        if max_x > 60:
            x_ticks.append(int(max_x))
        plt.xticks(x_ticks, fontsize=9)

        plt.axvline(x=5, color='#DD6B20', linestyle='--', alpha=0.6, linewidth=1.5)
        plt.axvline(x=20, color='#E53E3E', linestyle='--', alpha=0.6, linewidth=1.5)
        plt.axvline(x=60, color='#822727', linestyle='--', alpha=0.6, linewidth=1.5)
        text_y = len(locations) - 0.5 
        plt.text(2.5, len(locations)-0.5, 'Bajo', rotation=90, va='top', ha='center',
             color='#38A169', fontweight='bold', fontsize=9,
             bbox=dict(facecolor='white', alpha=0.5, pad=3, edgecolor='#38A169'))

        plt.text(12.5, len(locations)-0.5, 'Moderado', rotation=90, va='top', ha='center',
             color='#DD6B20', fontweight='bold', fontsize=9,
             bbox=dict(facecolor='white', alpha=0.5, pad=3, edgecolor='#DD6B20'))

        plt.text(40, len(locations)-0.5, 'Alto', rotation=90, va='top', ha='center',
             color='#E53E3E', fontweight='bold', fontsize=9,
             bbox=dict(facecolor='white', alpha=0.5, pad=3, edgecolor='#E53E3E'))

        if max_x > 60:
            plt.text(60 + (max_x-60)/2, len(locations)-0.5, 'Crítico', rotation=90, va='top', ha='center',
                 color='#822727', fontweight='bold', fontsize=9,
                 bbox=dict(facecolor='white', alpha=0.5, pad=3, edgecolor='#822727'))

        plt.yticks(fontsize=10, color='#2D3748')
        plt.gcf().set_facecolor('#FFFFFF')
        plt.gca().set_facecolor('#F7FAFC')
        plt.tight_layout(pad=3.0)

        img_buffer = io.BytesIO()
        plt.savefig(img_buffer, format='png', dpi=200, bbox_inches='tight')
        img_buffer.seek(0)

        img = Image(img_buffer)
        img.drawHeight = min(8 * inch, len(locations) * 0.9 * inch)
        img.drawWidth = 5.5 * inch
        elements.append(img)
        plt.close('all')

        elements.append(Spacer(1, 0.3 * inch))

        duration_table_data = [["Ubicación", "Duración Promedio", "Categoría"]]
        sorted_data = sorted(stats['duracion_promedio'], key=lambda x: x['duracion_promedio'], reverse=True)

        for item in sorted_data:
            location = item['nombre_posicion']
            duration = item['duracion_promedio']

            if duration < 5:
                category = "Bajo"
            elif duration < 20:
                category = "Moderado"
            elif duration < 60:
                category = "Alto"
            else:
                category = "Crítico"

            duration_table_data.append([
                location,
                f"{duration:.1f} segundos",
                category
            ])

        col_widths = [3 * inch, 2 * inch, 2 * inch]
        duration_table = Table(duration_table_data, colWidths=col_widths)

        duration_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A365D')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (2, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E0')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
        ]

        for i in range(1, len(duration_table_data)):
            if i % 2 == 0:
                duration_style.append(('BACKGROUND', (0, i), (1, i), colors.HexColor('#F7FAFC')))

            category = duration_table_data[i][2]

            if category == "Bajo":
                duration_style.append(('TEXTCOLOR', (2, i), (2, i), colors.HexColor('#38A169')))
                duration_style.append(('BACKGROUND', (2, i), (2, i), colors.HexColor('#F0FFF4')))
            elif category == "Moderado":
                duration_style.append(('TEXTCOLOR', (2, i), (2, i), colors.HexColor('#DD6B20')))
                duration_style.append(('BACKGROUND', (2, i), (2, i), colors.HexColor('#FFFAF0')))
            elif category == "Alto":
                duration_style.append(('TEXTCOLOR', (2, i), (2, i), colors.HexColor('#E53E3E')))
                duration_style.append(('BACKGROUND', (2, i), (2, i), colors.HexColor('#FFF5F5')))
            else:
                duration_style.append(('TEXTCOLOR', (2, i), (2, i), colors.HexColor('#822727')))
                duration_style.append(('BACKGROUND', (2, i), (2, i), colors.HexColor('#FFF0F0')))

            duration_style.append(('FONTNAME', (2, i), (2, i), 'Helvetica-Bold'))

        duration_table.setStyle(TableStyle(duration_style))
        elements.append(duration_table)

        elements.append(Spacer(1, 0.3 * inch))
        elements.append(Paragraph("Observaciones y Recomendaciones", section_style))

        has_critical = any(item['duracion_promedio'] >= 60 for item in stats['duracion_promedio'])
        has_high = any(20 <= item['duracion_promedio'] < 60 for item in stats['duracion_promedio'])

        recommendations = """
        <para>
        """

        if has_critical:
            recommendations += """
            <br/><br/><b>• Se detectaron ubicaciones con eventos de duración crítica</b> (más de 60 segundos).
            Estas ubicaciones requieren revisión prioritaria, ya que podrían indicar:
            <br/>  - Posibles problemas de seguridad persistentes
            <br/>  - Configuración incorrecta de sensibilidad en sensores
            <br/>  - Eventos anómalos que requieren atención inmediata
            """

        if has_high:
            recommendations += """
            <br/><br/><b>• Existen ubicaciones con eventos de duración alta</b> (entre 20 y 60 segundos).
            Se recomienda:
            <br/>  - Revisar patrones temporales de estos eventos
            <br/>  - Verificar si coinciden con horarios específicos o actividades planificadas
            <br/>  - Considerar ajustes moderados en la configuración de sensibilidad
            """

        recommendations += """
        <br/><br/>Se sugiere establecer un umbral de revisión periódica para ubicaciones que muestren
        consistentemente duraciones superiores a 20 segundos, y documentar las acciones tomadas para
        resolver cualquier problema identificado.
        </para>
        """

        elements.append(Paragraph(recommendations, normal_style))
        elements.append(Spacer(1, 0.3 * inch))
      
        
        # 1.3 Lista de eventos de larga duración (eventos críticos)
        if stats and 'eventos_largos' in stats and len(stats['eventos_largos']) > 0:
            elements.append(PageBreak())  # Nueva página para eventos largos
            elements.append(Paragraph("Eventos de Larga Duración", subtitle_style))
            
            # Texto explicativo
            long_events_text = """
            <para>Los eventos de larga duración pueden indicar situaciones que requieren atención prioritaria.
            La siguiente tabla muestra los eventos con duración superior a 20 segundos, ordenados de mayor a menor duración.</para>
            """
            elements.append(Paragraph(long_events_text, normal_style))
            elements.append(Spacer(1, 0.2*inch))
            
            # Cabeceras de la tabla
            long_events_data = [["Ubicación", "Fecha", "Hora", "Duración", "Estado"]]
            
            # Datos de eventos largos
            for event in stats['eventos_largos']:
                duration_str = ""
                duration = event['duracion_segundos']
                
                # Formatear duración para mejor legibilidad
                if duration >= 60:
                    minutes = int(duration // 60)
                    seconds = int(duration % 60)
                    duration_str = f"{minutes}m {seconds}s"
                else:
                    duration_str = f"{int(duration)}s"
                
                long_events_data.append([
                    event['nombre_posicion'],
                    event['fecha_evento'],
                    event['hora_evento'],
                    duration_str,
                    "✓ Revisado" if event['revisado'] else "⚠ Pendiente"
                ])
            
            # Crear tabla
            col_widths = [2.5*inch, 1*inch, 1*inch, 1*inch, 1.5*inch]
            long_events_table = Table(long_events_data, colWidths=col_widths, repeatRows=1)
            
            # Estilo de la tabla
            table_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#B3C1D1'))
            ]
            
            # Colorear filas según duración
            for i, event in enumerate(stats['eventos_largos'], 1):
                duration = event['duracion_segundos']
                
                if duration >= 60:
                    table_style.append(('BACKGROUND', (3, i), (3, i), colors.HexColor('#FFEBEE')))  # Rojo claro
                    table_style.append(('TEXTCOLOR', (3, i), (3, i), colors.HexColor('#D32F2F')))  # Rojo oscuro
                elif duration >= 20:
                    table_style.append(('BACKGROUND', (3, i), (3, i), colors.HexColor('#FFF8E1')))  # Amarillo claro
                    table_style.append(('TEXTCOLOR', (3, i), (3, i), colors.HexColor('#F57F17')))  # Naranja oscuro
                
                # Destacar eventos no revisados
                if not event['revisado']:
                    table_style.append(('TEXTCOLOR', (-1, i), (-1, i), colors.HexColor('#CC3300')))
                    table_style.append(('FONTNAME', (-1, i), (-1, i), 'Helvetica-Bold'))
                else:
                    table_style.append(('TEXTCOLOR', (-1, i), (-1, i), colors.HexColor('#006633')))
                
                # Filas alternadas
                if i % 2 == 0:
                    table_style.append(('BACKGROUND', (0, i), (-2, i), colors.HexColor('#F0F5FA')))
            
            long_events_table.setStyle(TableStyle(table_style))
            elements.append(long_events_table)
            
            # Añadir recomendaciones para eventos largos
            elements.append(Spacer(1, 0.3*inch))
            recommendations_title = "Recomendaciones para Eventos de Larga Duración"
            elements.append(Paragraph(recommendations_title, subtitle_style))
            
            recommendations_text = """
            <para>Basado en el análisis de los eventos de larga duración, se recomienda:
            <br/>• Verificar las ubicaciones con eventos de más de 60 segundos de duración.
            <br/>• Revisar la configuración de sensibilidad en las cámaras con alto promedio de duración.
            <br/>• Considerar ajustar los umbrales de detección para eventos repetitivos en las mismas ubicaciones.
            <br/>• Establecer protocolos de respuesta prioritaria para eventos críticos (>60s).
            </para>
            """
            elements.append(Paragraph(recommendations_text, normal_style))
        
    
    # Reporte tipo Actividad por Cámara
    elif report_type == "camera-activity":
        # Para actividad por cámara, enfocamos en distribución y tendencias
        elements.append(Paragraph("Análisis de Actividad por Cámara", title_style))
        elements.append(Spacer(1, 0.2*inch))
        elements.append(Paragraph(
        "Este reporte detalla la actividad detectada por cada cámara, resaltando "
        "la distribución de alertas, los periodos de mayor actividad y la tendencia semanal. "
        "Estos indicadores permiten evaluar el desempeño del sistema de vigilancia y la efectividad de cobertura.",
        normal_style))
        elements.append(Spacer(1, 0.3*inch))
        
        # Distribución por cámara como gráfico principal
        if charts_data and 'by_camera' in charts_data and len(charts_data['by_camera']['labels']) > 0:
            # Gráfico de pastel mejorado visualmente
            plt.figure(figsize=(7, 5))
            
            # Usar colores más atractivos
            color_palette = plt.cm.viridis(np.linspace(0, 1, len(charts_data['by_camera']['labels'])))
            
            # Si hay muchas cámaras, limitamos para mejor visualización
            max_cameras = 8
            if len(charts_data['by_camera']['labels']) > max_cameras:
                # Tomar las cámaras con más alertas
                top_cameras = sorted(zip(charts_data['by_camera']['labels'], 
                                        charts_data['by_camera']['data']),
                                    key=lambda x: x[1], reverse=True)
                
                # Separar top cámaras y agrupar el resto
                top_labels = [x[0] for x in top_cameras[:max_cameras-1]]
                top_data = [x[1] for x in top_cameras[:max_cameras-1]]
                
                # Crear categoría "Otras"
                other_data = sum([x[1] for x in top_cameras[max_cameras-1:]])
                
                # Juntar datos
                labels = top_labels + ["Otras"]
                data = top_data + [other_data]
                
                wedges, texts, autotexts = plt.pie(data, 
                                                  labels=None,
                                                  autopct='%1.1f%%',
                                                  startangle=90,
                                                  colors=color_palette[:max_cameras])
            else:
                wedges, texts, autotexts = plt.pie(charts_data['by_camera']['data'], 
                                                 labels=None,
                                                 autopct='%1.1f%%',
                                                 startangle=90,
                                                 colors=color_palette)
                labels = charts_data['by_camera']['labels']
                data = charts_data['by_camera']['data']
            
            # Personalizar look & feel
            plt.title('Distribución de Alertas por Cámara', fontsize=14, pad=20)
            plt.axis('equal')
            
            # Leyenda en una posición mejor
            plt.legend(wedges, labels, title="Cámaras", 
                      loc="center left", 
                      bbox_to_anchor=(1, 0.5),
                      fontsize=9)
            
            # Personalizar porcentajes
            for autotext in autotexts:
                autotext.set_fontsize(9)
                autotext.set_weight('bold')
            
            plt.tight_layout()
            
            # Guardar la gráfica
            img_buffer = io.BytesIO()
            plt.subplots_adjust(left=0.25) 
            plt.savefig(img_buffer, format='png', dpi=150)
            img_buffer.seek(0)
            
            # Crear imagen para el PDF
            img = Image(img_buffer)
            img.drawHeight = 4*inch
            img.drawWidth = 7*inch
            elements.append(img)
            plt.close('all')
            
            elements.append(Spacer(1, 0.3*inch))
        
        # Tabla de resumen por cámara
        if charts_data and 'by_camera' in charts_data and len(charts_data['by_camera']['labels']) > 0:
            camera_data = []
            
            # Cabecera
            camera_data.append(["Posición de Cámara", "Total Alertas", "% del Total"])
            
            # Calcular total
            total = sum(charts_data['by_camera']['data'])
            
            # Datos
            for i, cam in enumerate(charts_data['by_camera']['labels']):
                count = charts_data['by_camera']['data'][i]
                percentage = (count/total*100) if total > 0 else 0
                camera_data.append([
                    cam,
                    str(count),
                    f"{percentage:.1f}%"
                ])
            
            # Crear tabla
            camera_table = Table(camera_data, colWidths=[3*inch, 1.5*inch, 1.5*inch], repeatRows=1)
            
            # Estilo
            cam_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#B3C1D1'))
            ]
            
            # Colorear filas alternadas
            for i in range(1, len(camera_data)):
                if i % 2 == 0:
                    cam_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F0F5FA')))
            
            camera_table.setStyle(TableStyle(cam_style))
            elements.append(camera_table)
        
        # Incluir tendencia semanal si hay datos
        if charts_data and 'trend' in charts_data and len(charts_data['trend']['labels']) > 0:
            elements.append(Spacer(1, 0.4*inch))
            elements.append(Paragraph("Tendencia de Alertas - Últimas Semanas", subtitle_style))
            
            # Gráfico de línea con tendencia
            plt.figure(figsize=(7, 3))
            labels = charts_data['trend']['labels']
            data = charts_data['trend']['datasets'][0]['data']
            
            plt.plot(labels, data, marker='o', color='#0066CC', linewidth=2, markersize=8)
            
            # Añadir área bajo la curva
            plt.fill_between(labels, data, alpha=0.2, color='#0066CC')
            
            # Personalizar gráfico
            plt.title('Evolución de Alertas por Semana')
            plt.grid(True, linestyle='--', alpha=0.7)
            
            # Añadir valores sobre los puntos
            for i, v in enumerate(data):
                plt.text(i, v + max(data)*0.05, str(v), ha='center')
            
            plt.tight_layout()
            
            # Guardar la gráfica
            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format='png', dpi=150)
            img_buffer.seek(0)
            
            # Crear imagen para el PDF
            img = Image(img_buffer)
            img.drawHeight = 3*inch
            img.drawWidth = 7*inch
            elements.append(img)
            plt.close('all')
    
    # Reporte Completo
    else:  # report_type == "complete"
    # Incluimos información completa integrando tanto el reporte de cámaras como el de sensores y duración
    
    # Título principal y descripción del reporte completo
        elements.append(PageBreak())
        elements.append(Paragraph("REPORTE DE SEGURIDAD COMPLETO", title_style))
        elements.append(Spacer(1, 0.2*inch))
    
    # Texto introductorio explicando el reporte completo
    intro_text = """
    <para>Este reporte integral combina el análisis de actividad por cámara y la evaluación de duración de eventos, 
    proporcionando una visión completa del sistema de seguridad. Se incluyen métricas clave, tendencias 
    de actividad, análisis de duración de eventos y recomendaciones específicas para optimizar el rendimiento 
    del sistema y reforzar la seguridad.</para>
    """
    elements.append(Paragraph(intro_text, normal_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # 1. SECCIÓN DE RESUMEN ESTADÍSTICO CONSOLIDADO
    # Creamos un panel visualmente atractivo con estadísticas clave combinadas
    elements.append(Paragraph("1. Resumen Estadístico Consolidado", section_style))
    
    if stats:
        # Creamos tabla para métricas principales con diseño moderno
        metrics_data = [
            ["Métrica", "Valor", "Indicador Visual"],
            ["Total de Alertas", f"{stats['total_alerts']}", "■" * min(20, int(stats['total_alerts']/10 + 1))],
            ["Alertas Recientes (24h)", f"{stats['recent_alerts']}", "■" * min(20, int(stats['recent_alerts']/3 + 1))],
            ["Alertas Pendientes", f"{stats['total_alerts'] - stats['reviewed_alerts']}", "■" * min(20, int((stats['total_alerts'] - stats['reviewed_alerts'])/5 + 1))],
            ["Alertas Críticas (>60s)", f"{stats['duracion_categorias']['critico']}", "■" * min(20, int(stats['duracion_categorias']['critico']*2 + 1))],
        ]
        
        # Calculamos porcentaje de revisión para el indicador de estado
        if stats['total_alerts'] > 0:
            review_percentage = (stats['reviewed_alerts'] / stats['total_alerts']) * 100
            review_status = f"{review_percentage:.1f}% Revisadas"
        else:
            review_status = "No hay alertas"
        
        metrics_table = Table(metrics_data, colWidths=[2.5*inch, 1.5*inch, 3*inch])
        metrics_style = [
            # Cabecera
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A365D')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('ALIGN', (2, 0), (2, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E0')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
        ]
        
        # Colorear las filas para mejor visualización
        for i in range(1, len(metrics_data)):
            if i % 2 == 0:
                metrics_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F7FAFC')))
                
            # Destacar las alertas críticas con color
            if i == 4:  # Fila de alertas críticas
                metrics_style.append(('TEXTCOLOR', (1, i), (1, i), colors.HexColor('#E53E3E')))
                metrics_style.append(('FONTNAME', (1, i), (1, i), 'Helvetica-Bold'))
                metrics_style.append(('BACKGROUND', (2, i), (2, i), colors.HexColor('#FFF5F5')))
        
        metrics_table.setStyle(TableStyle(metrics_style))
        elements.append(metrics_table)
        
        # Añadir un resumen textual con indicadores de salud del sistema
        elements.append(Spacer(1, 0.2*inch))
        
        # Determinar el estado del sistema basado en métricas
        system_status = "NORMAL"
        status_color = "#38A169"  # Verde por defecto
        
        if stats['duracion_categorias']['critico'] > 5 or (stats['total_alerts'] > 0 and (stats['total_alerts'] - stats['reviewed_alerts'])/stats['total_alerts'] > 0.3):
            system_status = "REQUIERE ATENCIÓN"
            status_color = "#E53E3E"  # Rojo
        elif stats['duracion_categorias']['alto'] > 10 or (stats['total_alerts'] > 0 and (stats['total_alerts'] - stats['reviewed_alerts'])/stats['total_alerts'] > 0.2):
            system_status = "PRECAUCIÓN"
            status_color = "#DD6B20"  # Naranja
        
        # Panel de estado del sistema
        status_table = Table([["Estado del Sistema:", system_status]], 
                             colWidths=[2*inch, 5*inch])
        status_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'RIGHT'),
            ('ALIGN', (1, 0), (1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (1, 0), 12),
            ('TEXTCOLOR', (1, 0), (1, 0), colors.HexColor(status_color)),
            ('FONTSIZE', (1, 0), (1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (1, 0), 10),
            ('TOPPADDING', (0, 0), (1, 0), 10),
        ]))
        elements.append(status_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # 2. ANÁLISIS DE DISTRIBUCIÓN DE ACTIVIDAD
    elements.append(Paragraph("2. Análisis de Distribución de Actividad", section_style))
    
    # 2.1 Distribución por día y hora
    if charts_data and ('by_day' in charts_data or 'by_hour' in charts_data):
        # Crear gráficos en la misma fila si es posible
        day_hour_data = []
        
        # Si tenemos datos por día
        if 'by_day' in charts_data:
            plt.figure(figsize=(5, 3.5))
            bars = plt.bar(charts_data['by_day']['labels'], charts_data['by_day']['data'], 
                       color=plt.cm.Blues(np.linspace(0.6, 0.9, len(charts_data['by_day']['labels']))))
            
            # Añadir etiquetas sobre las barras
            for bar in bars:
                height = bar.get_height()
                plt.text(bar.get_x() + bar.get_width()/2., height + 0.5,
                       '%d' % int(height), ha='center', va='bottom', fontweight='bold')
            
            plt.title('Distribución por Día de la Semana', fontsize=12, fontweight='bold')
            plt.grid(axis='y', alpha=0.3, linestyle='--')
            plt.ylabel('Número de Alertas', fontsize=10)
            plt.ylim(0, max(charts_data['by_day']['data']) * 1.2)
            
            ax = plt.gca()
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            ax.set_facecolor('#F8FAFC')
            
            plt.tight_layout()
            
            day_buffer = io.BytesIO()
            plt.savefig(day_buffer, format='png', dpi=150, bbox_inches='tight')
            day_buffer.seek(0)
            day_img = Image(day_buffer)
            day_img.drawHeight = 2.8*inch
            day_img.drawWidth = 3.5*inch
            plt.close('all')
            
            if 'by_hour' in charts_data:
                # Si tenemos también datos por hora, usaremos una tabla para organizar los gráficos
                plt.figure(figsize=(5, 3.5))
                
                # Crear gráfico por hora con estilo coherente
                plt.plot(charts_data['by_hour']['labels'], charts_data['by_hour']['data'], 
                       marker='o', color='#2B6CB0', linewidth=2)
                
                # Añadir área bajo la curva
                plt.fill_between(charts_data['by_hour']['labels'], charts_data['by_hour']['data'], 
                               alpha=0.2, color='#2B6CB0')
                
                plt.title('Distribución por Hora del Día', fontsize=12, fontweight='bold')
                plt.grid(True, alpha=0.3, linestyle='--')
                plt.ylabel('Número de Alertas', fontsize=10)
                plt.xlabel('Hora', fontsize=10)
                
                # Mostrar solo algunas horas para evitar sobrecarga visual
                plt.xticks([0, 6, 12, 18, 23])
                
                ax = plt.gca()
                ax.spines['top'].set_visible(False)
                ax.spines['right'].set_visible(False)
                ax.set_facecolor('#F8FAFC')
                
                plt.tight_layout()
                
                hour_buffer = io.BytesIO()
                plt.savefig(hour_buffer, format='png', dpi=150, bbox_inches='tight')
                hour_buffer.seek(0)
                hour_img = Image(hour_buffer)
                hour_img.drawHeight = 2.8*inch
                hour_img.drawWidth = 3.5*inch
                plt.close('all')
                
                # Crear tabla para organizar los gráficos en dos columnas
                graphics_table_data = [[day_img, hour_img]]
                graphics_table = Table(graphics_table_data, colWidths=[3.5*inch, 3.5*inch])
                graphics_table.setStyle(TableStyle([
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ]))
                elements.append(graphics_table)
            else:
                # Si solo tenemos datos por día
                elements.append(day_img)
        
        # Espacio después de los gráficos
        elements.append(Spacer(1, 0.3*inch))
    
    # 2.2 Distribución por Cámara
    if charts_data and 'by_camera' in charts_data and len(charts_data['by_camera']['labels']) > 0:
        elements.append(Paragraph("Distribución de Alertas por Cámara", subtitle_style))
        
        # Crear gráfico de pastel mejorado
        plt.figure(figsize=(6, 4.5))
        
        # Colores más atractivos
        color_palette = plt.cm.viridis(np.linspace(0, 1, len(charts_data['by_camera']['labels'])))
        
        # Si hay muchas cámaras, limitamos para mejor visualización
        max_cameras = 8
        if len(charts_data['by_camera']['labels']) > max_cameras:
            # Tomar las cámaras con más alertas
            top_cameras = sorted(zip(charts_data['by_camera']['labels'], 
                                    charts_data['by_camera']['data']),
                                key=lambda x: x[1], reverse=True)
            
            # Separar top cámaras y agrupar el resto
            top_labels = [x[0] for x in top_cameras[:max_cameras-1]]
            top_data = [x[1] for x in top_cameras[:max_cameras-1]]
            
            # Crear categoría "Otras"
            other_data = sum([x[1] for x in top_cameras[max_cameras-1:]])
            
            # Juntar datos
            labels = top_labels + ["Otras"]
            data = top_data + [other_data]
            
            wedges, texts, autotexts = plt.pie(data, 
                                             labels=None,
                                             autopct='%1.1f%%',
                                             startangle=90,
                                             colors=color_palette[:max_cameras],
                                             shadow=True)
        else:
            wedges, texts, autotexts = plt.pie(charts_data['by_camera']['data'], 
                                             labels=None,
                                             autopct='%1.1f%%',
                                             startangle=90,
                                             colors=color_palette,
                                             shadow=True)
            labels = charts_data['by_camera']['labels']
            data = charts_data['by_camera']['data']
        
        # Personalizar look & feel
        plt.title('Distribución de Alertas por Cámara', fontsize=14, pad=20, fontweight='bold')
        plt.axis('equal')
        
        # Leyenda en una posición mejor
        plt.legend(wedges, labels, title="Cámaras", 
                  loc="center left", 
                  bbox_to_anchor=(1, 0.5),
                  fontsize=9)
        
        # Personalizar porcentajes
        for autotext in autotexts:
            autotext.set_fontsize(9)
            autotext.set_weight('bold')
            autotext.set_color('white')
        
        plt.tight_layout()
        
        # Guardar la gráfica
        img_buffer = io.BytesIO()
        plt.savefig(img_buffer, format='png', dpi=150)
        img_buffer.seek(0)
        
        # Crear imagen para el PDF
        img = Image(img_buffer)
        img.drawHeight = 4*inch
        img.drawWidth = 7*inch
        elements.append(img)
        plt.close('all')
        
        elements.append(Spacer(1, 0.3*inch))
        
        # Tabla resumen por cámara
        elements.append(Paragraph("Resumen de Actividad por Cámara", subtitle_style))
        
        camera_data = []
        
        # Cabecera
        camera_data.append(["Posición de Cámara", "Total Alertas", "% del Total", "Tendencia"])
        
        # Calcular total
        total = sum(charts_data['by_camera']['data'])
        
        # Simulamos datos de tendencia (en un sistema real vendríamos de la base de datos)
        trends = ["↑", "↓", "→", "↑", "↓", "→", "↑", "↓", "→", "↑", "↓", "→"]  # Tendencias de ejemplo
        
        # Datos
        for i, cam in enumerate(charts_data['by_camera']['labels']):
            count = charts_data['by_camera']['data'][i]
            percentage = (count/total*100) if total > 0 else 0
            
            # Asignar tendencia (en un sistema real esto vendría de comparar con períodos anteriores)
            trend = trends[i % len(trends)]
            trend_color = "#38A169" if trend == "↑" else "#E53E3E" if trend == "↓" else "#718096"
            
            camera_data.append([
                cam,
                str(count),
                f"{percentage:.1f}%",
                trend
            ])
        
        # Crear tabla
        camera_table = Table(camera_data, colWidths=[3*inch, 1.3*inch, 1.3*inch, 1.4*inch], repeatRows=1)
        
        # Estilo
        cam_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A365D')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E0')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]
        
        # Colorear filas alternadas y añadir colores a las tendencias
        for i in range(1, len(camera_data)):
            if i % 2 == 0:
                cam_style.append(('BACKGROUND', (0, i), (-2, i), colors.HexColor('#F7FAFC')))
            
            # Color de tendencia
            trend = camera_data[i][3]
            trend_color = "#38A169" if trend == "↑" else "#E53E3E" if trend == "↓" else "#718096"
            cam_style.append(('TEXTCOLOR', (3, i), (3, i), colors.HexColor(trend_color)))
            cam_style.append(('FONTNAME', (3, i), (3, i), 'Helvetica-Bold'))
            cam_style.append(('FONTSIZE', (3, i), (3, i), 14))
        
        camera_table.setStyle(TableStyle(cam_style))
        elements.append(camera_table)
        
        elements.append(Spacer(1, 0.5*inch))
    
    # 3. ANÁLISIS DE DURACIÓN DE EVENTOS
    elements.append(Paragraph("3. Análisis de Duración de Eventos", section_style))
    
    if stats and 'duracion_categorias' in stats:
        # Texto explicativo mejorado con iconos visuales
        duration_explanation = """
        <para>El análisis de duración clasifica los eventos en categorías según el tiempo que permanecen activos:
        <br/><b>• <font color="#66BB6A">Bajo:</font></b> Menos de 5 segundos - Eventos breves, generalmente falsos positivos.
        <br/><b>• <font color="#FFC107">Moderado:</font></b> Entre 5 y 20 segundos - Duración normal para tránsito regular.
        <br/><b>• <font color="#FF9800">Alto:</font></b> Entre 20 y 60 segundos - Actividad prolongada que merece revisión.
        <br/><b>• <font color="#F44336">Crítico:</font></b> Más de 60 segundos - Posible situación de riesgo que requiere atención inmediata.
        </para>
        """
        elements.append(Paragraph(duration_explanation, normal_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Gráfico de distribución por categorías de duración MEJORADO
        plt.figure(figsize=(7, 4.5))
        categories = ['Bajo (<5s)', 'Moderado (5-20s)', 'Alto (20-60s)', 'Crítico (>60s)']
        values = [
            stats['duracion_categorias']['bajo'],
            stats['duracion_categorias']['moderado'],
            stats['duracion_categorias']['alto'],
            stats['duracion_categorias']['critico']
        ]
        
        # Colores para indicar severidad con un toque más vibrante pero profesional
        bar_colors = ['#66BB6A', '#FFC107', '#FF9800', '#F44336']
        
        # Crear gráfico de barras con estilo mejorado
        ax = plt.subplot(111)
        bars = plt.bar(categories, values, color=bar_colors, width=0.65)
        
        # Añadir valores sobre las barras
        for bar in bars:
            height = bar.get_height()
            plt.text(bar.get_x() + bar.get_width()/2., height + 0.8,
                   f'{int(height)}', ha='center', va='bottom', fontweight='bold', fontsize=12)
        
        # Mejorar diseño del gráfico
        plt.title('Distribución de Eventos por Duración', fontsize=16, pad=20, fontweight='bold')
        plt.ylabel('Número de eventos', fontsize=12)
        plt.grid(axis='y', linestyle='--', alpha=0.3)
        
        # Eliminar bordes innecesarios
        for spine in ax.spines.values():
            spine.set_visible(False)
        
        # Añadir colores de fondo para mejor visualización
        ax.set_facecolor('#F7FAFC')  # Fondo gris muy claro
        plt.gcf().set_facecolor('#FFFFFF')
        
        # Ajustar ticks del eje Y para mejor legibilidad
        plt.yticks(fontsize=8)
        plt.xticks(fontsize=9, fontweight='bold')
        
        # Añadir valor total como anotación
        total = sum(values)
        plt.annotate(f'Total: {total} eventos', 
                   xy=(0.5, 0.97),
                   xycoords='axes fraction',
                   fontsize=12,
                   fontweight='bold',
                   ha='center',
                   va='top',
                   bbox=dict(boxstyle="round,pad=0.3", fc="#EBF8FF", ec="#4299E1", alpha=0.8))
        
        plt.tight_layout()
        
        # Guardar gráfica
        img_buffer = io.BytesIO()
        plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
        img_buffer.seek(0)
        
        # Crear imagen para el PDF
        img = Image(img_buffer)
        img.drawHeight = 3.5*inch
        img.drawWidth = 7*inch
        elements.append(img)
        plt.close('all')
        
        elements.append(Spacer(1, 0.3*inch))
        
        # Tabla resumen con porcentajes - MEJORADA
        total_events = sum(values)
        duration_data = [
            ["Categoría", "Eventos", "Porcentaje", "Nivel de Atención"]
        ]
        
        # Definiciones de niveles de atención para cada categoría
        attention_levels = [
            "Bajo - Revisión opcional",
            "Moderado - Revisión recomendada",
            "Alto - Requiere revisión",
            "Crítico - Atención inmediata"
        ]
        
        for i, cat in enumerate(categories):
            count = values[i]
            percentage = (count/total_events*100) if total_events > 0 else 0
            duration_data.append([
                cat,
                str(count),
                f"{percentage:.1f}%",
                attention_levels[i]
            ])
        
        # Añadir fila de totales
        duration_data.append([
            "TOTAL",
            str(total_events),
            "100.0%",
            ""
        ])
        
        # Crear tabla con mejor diseño
        duration_table = Table(duration_data, colWidths=[1.5*inch, 1*inch, 1*inch, 3.5*inch])
        
        # Estilo de la tabla más moderno
        duration_style = [
            # Cabecera
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A365D')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (2, -1), 'CENTER'),
            ('ALIGN', (3, 0), (3, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E0')),
            # Configuraciones generales
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 1), (-1, -2), 8),
            ('TOPPADDING', (0, 1), (-1, -2), 8),
        ]
        
        # Mejorar colores para categorías - contraste mejorado
        color_hex = ['#E8F5E9', '#FFF8E1', '#FFF3E0', '#FFEBEE']  # Fondos claros para mejor legibilidad
        text_colors = ['#2E7D32', '#F57F17', '#E65100', '#B71C1C']  # Textos con contraste
        
        for i in range(4):
            row = i + 1
            duration_style.append(('BACKGROUND', (0, row), (0, row), colors.HexColor(color_hex[i])))
            duration_style.append(('TEXTCOLOR', (0, row), (0, row), colors.HexColor(text_colors[i])))
            duration_style.append(('FONTNAME', (0, row), (0, row), 'Helvetica-Bold'))
        
        # Fila de totales con estilo distintivo
        duration_style.append(('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#EDF2F7')))
        duration_style.append(('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'))
        duration_style.append(('LINEBELOW', (0, -2), (-1, -2), 1, colors.HexColor('#4A5568')))
        duration_style.append(('TOPPADDING', (0, -1), (-1, -1), 10))
        duration_style.append(('BOTTOMPADDING', (0, -1), (-1, -1), 10))
        
        duration_table.setStyle(TableStyle(duration_style))
        elements.append(duration_table)
        
        # 3.1 Duración promedio por ubicación
        if stats and 'duracion_promedio' in stats and len(stats['duracion_promedio']) > 0:
            elements.append(Paragraph("Duración Promedio por Ubicación", subtitle_style))
            
            avg_duration_text = """
            <para>El análisis de duración promedio por ubicación permite identificar patrones espaciales en la activación
            de sensores. Las ubicaciones con valores más altos pueden requerir revisión de configuración o indicar
            situaciones que demandan atención prioritaria.</para>
            """
            elements.append(Paragraph(avg_duration_text, normal_style))
            elements.append(Spacer(1, 0.2 * inch))
            
            locations = [item['nombre_posicion'] for item in stats['duracion_promedio']]
            avg_durations = [item['duracion_promedio'] for item in stats['duracion_promedio']]
            
            if len(locations) > 10:
                sorted_data = sorted(zip(locations, avg_durations), key=lambda x: x[1], reverse=True)
                locations = [item[0] for item in sorted_data[:10]]
                avg_durations = [item[1] for item in sorted_data[:10]]
                note = "Nota: Se muestran las 10 ubicaciones con mayor duración promedio."
                elements.append(Paragraph(note, ParagraphStyle(
                'Note',
                parent=normal_style,
                fontSize=9,
                textColor=colors.HexColor('#718096'),
                fontName='Helvetica-Oblique'
                )))
            
            plt.figure(figsize=(5.5, max(4, len(locations) * 0.6)), dpi=150)
            plt.style.use('seaborn-v0_8-whitegrid')
            
            bar_colors = []
            for duration in avg_durations:
                if duration < 5:
                    bar_colors.append('#38A169')
                elif duration < 20:
                    bar_colors.append('#DD6B20')
                elif duration < 60:
                    bar_colors.append('#E53E3E')
                else:
                    bar_colors.append('#822727')
            
            bars = plt.barh(locations, avg_durations, color=bar_colors, height=0.7,
                           edgecolor='white', linewidth=0.7)
            # Añadir valores sobre las barras
            for bar in bars:
                height = bar.get_height()
                plt.text(bar.get_x() + bar.get_width()/2., height + 0.8,
                       f'{int(height)}', ha='center', va='bottom', fontweight='bold', fontsize=12)
            
            # Mejorar diseño del gráfico
            plt.title('Distribución de Eventos por Duración', fontsize=16, pad=20, fontweight='bold')
            plt.ylabel('Número de eventos', fontsize=12)
            plt.grid(axis='y', linestyle='--', alpha=0.3)
            
            # Eliminar bordes innecesarios
            for spine in ax.spines.values():
                spine.set_visible(False)
            
            # Añadir colores de fondo para mejor visualización
            ax.set_facecolor('#F7FAFC')  # Fondo gris muy claro
            plt.gcf().set_facecolor('#FFFFFF')
            
            # Ajustar ticks del eje Y para mejor legibilidad
            plt.yticks(fontsize=8)
            plt.xticks(fontsize=9, fontweight='bold')
            
            # Añadir valor total como anotación
            total = sum(values)
            plt.annotate(f'Total: {total} eventos', 
                       xy=(0.5, 0.97),
                       xycoords='axes fraction',
                       fontsize=12,
                       fontweight='bold',
                       ha='center',
                       va='top',
                       bbox=dict(boxstyle="round,pad=0.3", fc="#EBF8FF", ec="#4299E1", alpha=0.8))
            
            plt.tight_layout()
            
            # Guardar gráfica
            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
            img_buffer.seek(0)
            
            # Crear imagen para el PDF
            img = Image(img_buffer)
            img.drawHeight = 3.5*inch
            img.drawWidth = 7*inch
            elements.append(img)
            plt.close('all')
            
            elements.append(Spacer(1, 0.3*inch))
            

        elements.append(Spacer(1, 0.3 * inch))
        elements.append(Paragraph("Observaciones y Recomendaciones", section_style))

        has_critical = any(item['duracion_promedio'] >= 60 for item in stats['duracion_promedio'])
        has_high = any(20 <= item['duracion_promedio'] < 60 for item in stats['duracion_promedio'])

        recommendations = """
        <para>
        """
        if has_critical:
            recommendations += """
            <br/><br/><b>• Se detectaron ubicaciones con eventos de duración crítica</b> (más de 60 segundos).
            Estas ubicaciones requieren revisión prioritaria, ya que podrían indicar:
            <br/>  - Posibles problemas de seguridad persistentes
            <br/>  - Configuración incorrecta de sensibilidad en sensores
            <br/>  - Eventos anómalos que requieren atención inmediata
            """

        if has_high:
            recommendations += """
            <br/><br/><b>• Existen ubicaciones con eventos de duración alta</b> (entre 20 y 60 segundos).
            Se recomienda:
            <br/>  - Revisar patrones temporales de estos eventos
            <br/>  - Verificar si coinciden con horarios específicos o actividades planificadas
            <br/>  - Considerar ajustes moderados en la configuración de sensibilidad
            """

        recommendations += """
        <br/><br/>Se sugiere establecer un umbral de revisión periódica para ubicaciones que muestren
        consistentemente duraciones superiores a 20 segundos, y documentar las acciones tomadas para
        resolver cualquier problema identificado.
        </para>
        """

        elements.append(Paragraph(recommendations, normal_style))
        elements.append(Spacer(1, 0.3 * inch))
      
        
        # 1.3 Lista de eventos de larga duración (eventos críticos)
        if stats and 'eventos_largos' in stats and len(stats['eventos_largos']) > 0:
            elements.append(Paragraph("Eventos de Larga Duración", subtitle_style))
            
            # Texto explicativo
            long_events_text = """
            <para>Los eventos de larga duración pueden indicar situaciones que requieren atención prioritaria.
            La siguiente tabla muestra los eventos con duración superior a 20 segundos, ordenados de mayor a menor duración.</para>
            """
            elements.append(Paragraph(long_events_text, normal_style))
            elements.append(Spacer(1, 0.2*inch))
            
            # Cabeceras de la tabla
            long_events_data = [["Ubicación", "Fecha", "Hora", "Duración", "Estado"]]
            
            # Datos de eventos largos
            for event in stats['eventos_largos']:
                duration_str = ""
                duration = event['duracion_segundos']
                
                # Formatear duración para mejor legibilidad
                if duration >= 60:
                    minutes = int(duration // 60)
                    seconds = int(duration % 60)
                    duration_str = f"{minutes}m {seconds}s"
                else:
                    duration_str = f"{int(duration)}s"
                
                long_events_data.append([
                    event['nombre_posicion'],
                    event['fecha_evento'],
                    event['hora_evento'],
                    duration_str,
                    "✓ Revisado" if event['revisado'] else "⚠ Pendiente"
                ])
            
            # Crear tabla
            col_widths = [2.5*inch, 1*inch, 1*inch, 1*inch, 1.5*inch]
            long_events_table = Table(long_events_data, colWidths=col_widths, repeatRows=1)
            
            # Estilo de la tabla
            table_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#B3C1D1'))
            ]
            
            # Colorear filas según duración
            for i, event in enumerate(stats['eventos_largos'], 1):
                duration = event['duracion_segundos']
                
                if duration >= 60:
                    table_style.append(('BACKGROUND', (3, i), (3, i), colors.HexColor('#FFEBEE')))  # Rojo claro
                    table_style.append(('TEXTCOLOR', (3, i), (3, i), colors.HexColor('#D32F2F')))  # Rojo oscuro
                elif duration >= 20:
                    table_style.append(('BACKGROUND', (3, i), (3, i), colors.HexColor('#FFF8E1')))  # Amarillo claro
                    table_style.append(('TEXTCOLOR', (3, i), (3, i), colors.HexColor('#F57F17')))  # Naranja oscuro
                
                # Destacar eventos no revisados
                if not event['revisado']:
                    table_style.append(('TEXTCOLOR', (-1, i), (-1, i), colors.HexColor('#CC3300')))
                    table_style.append(('FONTNAME', (-1, i), (-1, i), 'Helvetica-Bold'))
                else:
                    table_style.append(('TEXTCOLOR', (-1, i), (-1, i), colors.HexColor('#006633')))
                
                # Filas alternadas
                if i % 2 == 0:
                    table_style.append(('BACKGROUND', (0, i), (-2, i), colors.HexColor('#F0F5FA')))
            
            long_events_table.setStyle(TableStyle(table_style))
            elements.append(long_events_table)
            
            # Añadir recomendaciones para eventos largos
            elements.append(Spacer(1, 0.3*inch))
            recommendations_title = "Recomendaciones para Eventos de Larga Duración"
            elements.append(Paragraph(recommendations_title, subtitle_style))
            
            recommendations_text = """
            <para>Basado en el análisis de los eventos de larga duración, se recomienda:
            <br/>• Verificar las ubicaciones con eventos de más de 60 segundos de duración.
            <br/>• Revisar la configuración de sensibilidad en las cámaras con alto promedio de duración.
            <br/>• Considerar ajustar los umbrales de detección para eventos repetitivos en las mismas ubicaciones.
            <br/>• Establecer protocolos de respuesta prioritaria para eventos críticos (>60s).
            </para>
            """
            elements.append(Paragraph(recommendations_text, normal_style))
    elements.append(Spacer(1, 0.5*inch))
    footer_text = f"Sistema de Monitoreo de Seguridad | Reporte generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M:%S')}"
    elements.append(Paragraph(footer_text, ParagraphStyle('footer', parent=styles['Italic'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER)))
    # Construir el PDF
    doc.build(elements)
    return file_path

def generate_excel_report(title, user_info, stats, charts_data, alerts_data, report_type, date_range, start_date, end_date, filename):
    import os
    import pandas as pd
    import numpy as np
    import matplotlib.pyplot as plt
    import io
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference, Series, LineChart
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
    
    # Definir las funciones de estilo faltantes
    def apply_header_style(cell):
        """Aplica el estilo de encabezado a una celda"""
        cell.font = Font(name='Calibri', size=12, color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="4299E1")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin', color="CBD5E0"),
            right=Side(style='thin', color="CBD5E0"),
            top=Side(style='thin', color="CBD5E0"),
            bottom=Side(style='thin', color="CBD5E0")
        )
    
    def apply_subheader_style(cell):
        """Aplica el estilo de subencabezado a una celda"""
        cell.font = Font(name='Calibri', size=11, color="2A4365", bold=True)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = Border(
            bottom=Side(style='thin', color="CBD5E0")
        )
    
    def apply_normal_style(cell):
        """Aplica el estilo normal a una celda"""
        cell.font = Font(name='Calibri', size=11)
        cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Función auxiliar para obtener un nombre de archivo único
    def get_unique_filename(base_path, extension):
        """Obtiene un nombre de archivo único añadiendo un sufijo numérico si es necesario"""
        if not os.path.exists(f"{base_path}.{extension}"):
            return f"{base_path}.{extension}"
        
        counter = 1
        while os.path.exists(f"{base_path}_{counter}.{extension}"):
            counter += 1
        
        return f"{base_path}_{counter}.{extension}"
    
    # Crear directorio para reportes si no existe
    reports_dir = os.path.join(app.static_folder, 'reports')
    if not os.path.exists(reports_dir):
        os.makedirs(reports_dir)
    
    # Obtener una ruta de archivo única
    base_path = os.path.join(reports_dir, filename)
    file_path = get_unique_filename(base_path, 'xlsx')
    
    # Crear un nuevo libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    
    # Definir colores y estilos para el reporte
    title_fill = PatternFill("solid", fgColor="1A365D")
    title_font = Font(name='Calibri', size=14, color="FFFFFF", bold=True)
    subtitle_font = Font(name='Calibri', size=12, color="2A4365", bold=True)
    header_fill = PatternFill("solid", fgColor="4299E1")
    header_font = Font(name='Calibri', size=11, color="FFFFFF", bold=True)
    normal_font = Font(name='Calibri', size=11)
    centered_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color="CBD5E0"),
        right=Side(style='thin', color="CBD5E0"),
        top=Side(style='thin', color="CBD5E0"),
        bottom=Side(style='thin', color="CBD5E0")
    )
    
    # --- ENCABEZADO DEL REPORTE ---
    # Título
    ws.merge_cells('A1:H1')
    ws['A1'] = title
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = centered_align
    
    # Subtítulo según tipo de reporte
    if report_type == "cameras-analysis":
        subtitle = "ANÁLISIS DE CÁMARAS"
    elif report_type == "sensors-duration":
        subtitle = "ANÁLISIS DE SENSORES Y DURACIÓN"
    elif report_type == "complete-integrated":
        subtitle = "REPORTE COMPLETO INTEGRADO"
    else:
        subtitle = "REPORTE GENERAL"
    
    ws.merge_cells('A2:H2')
    ws['A2'] = subtitle
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = centered_align
    
    # Fecha de generación
    ws.merge_cells('A3:H3')
    ws['A3'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y - %H:%M:%S')}"
    ws['A3'].font = Font(name='Calibri', size=10, italic=True)
    ws['A3'].alignment = centered_align
    
    # --- INFORMACIÓN DEL USUARIO ---
    ws['A5'] = "INFORMACIÓN DEL USUARIO"
    ws['A5'].font = subtitle_font
    
    user_headers = ["Nombre:", "Usuario:"]
    user_values = [f"{user_info['nombres']} {user_info['apellidos']}", user_info['nombre_usuario']]
    
    for i, (header, value) in enumerate(zip(user_headers, user_values)):
        row = 6 + i
        ws[f'A{row}'] = header
        ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'B{row}'] = value
        ws[f'B{row}'].font = normal_font
    
    # Aplicar bordes a la sección de usuario
    for row in range(6, 8):
        for col in range(1, 3):
            ws.cell(row=row, column=col).border = thin_border
    
    # --- PERÍODO DEL REPORTE ---
    ws['A9'] = "PERÍODO ANALIZADO:"
    ws['A9'].font = subtitle_font
    
    if date_range == 'day':
        period_text = "Último día"
    elif date_range == 'week':
        period_text = "Última semana"
    elif date_range == 'month':
        period_text = "Último mes"
    elif date_range == 'custom' and start_date and end_date:
        period_text = f"Del {start_date} al {end_date}"
    else:
        period_text = "Período personalizado"
    
    ws['B9'] = period_text
    ws['B9'].font = normal_font
    
    # Si el reporte es de tipo "sensors-duration", generar hojas específicas
    if report_type == "sensors-duration":
        # --- HOJA: ANÁLISIS DE DURACIÓN ---
        ws_duration = wb.create_sheet("Análisis de Duración")
        
        # Título
        ws_duration.merge_cells('A1:F1')
        ws_duration['A1'] = "ANÁLISIS DE DURACIÓN DE EVENTOS"
        ws_duration['A1'].font = title_font
        ws_duration['A1'].fill = title_fill
        ws_duration['A1'].alignment = centered_align
        
        # Encabezados para las categorías de duración
        duration_headers = ["Categoría", "Rango", "Eventos", "Porcentaje", "Nivel de Atención"]
        for col, header in enumerate(duration_headers, 1):
            ws_duration.cell(row=3, column=col).value = header
            ws_duration.cell(row=3, column=col).font = header_font
            ws_duration.cell(row=3, column=col).fill = header_fill
            ws_duration.cell(row=3, column=col).alignment = centered_align
            ws_duration.cell(row=3, column=col).border = thin_border
        
        # Ancho de columnas
        ws_duration.column_dimensions['A'].width = 15
        ws_duration.column_dimensions['B'].width = 15
        ws_duration.column_dimensions['C'].width = 10
        ws_duration.column_dimensions['D'].width = 12
        ws_duration.column_dimensions['E'].width = 30
        
        # Datos de duración
        if stats and 'duracion_categorias' in stats:
            duration_data = [
                ["Bajo", "< 5 segundos", stats['duracion_categorias']['bajo'], "", "Bajo - Revisión opcional"],
                ["Moderado", "5-20 segundos", stats['duracion_categorias']['moderado'], "", "Moderado - Revisión recomendada"],
                ["Alto", "20-60 segundos", stats['duracion_categorias']['alto'], "", "Alto - Requiere revisión"],
                ["Crítico", "> 60 segundos", stats['duracion_categorias']['critico'], "", "Crítico - Atención inmediata"]
            ]
            
            # Calcular porcentajes
            total_events = sum(stats['duracion_categorias'][cat] for cat in ['bajo', 'moderado', 'alto', 'critico'])
            
            # Colores para las categorías
            category_fills = {
                "Bajo": PatternFill("solid", fgColor="66BB6A"),
                "Moderado": PatternFill("solid", fgColor="FFC107"),
                "Alto": PatternFill("solid", fgColor="FF9800"),
                "Crítico": PatternFill("solid", fgColor="F44336")
            }
            
            # Rellenar datos en la hoja
            for i, row_data in enumerate(duration_data):
                category = row_data[0]
                count = row_data[2]
                percentage = (count/total_events*100) if total_events > 0 else 0
                row_data[3] = f"{percentage:.1f}%"
                
                row = 4 + i
                for col, value in enumerate(row_data, 1):
                    cell = ws_duration.cell(row=row, column=col)
                    cell.value = value
                    cell.font = normal_font
                    cell.border = thin_border
                    
                    # Aplicar alineación
                    if col in [1, 2, 5]:
                        cell.alignment = left_align
                    else:
                        cell.alignment = centered_align
                
                # Colorear la celda de categoría
                ws_duration.cell(row=row, column=1).fill = category_fills[category]
                if category in ["Alto", "Crítico"]:
                    ws_duration.cell(row=row, column=1).font = Font(name='Calibri', size=11, color="FFFFFF", bold=True)
            
            # Añadir fila de total
            row = 8
            ws_duration.cell(row=row, column=1).value = "TOTAL"
            ws_duration.cell(row=row, column=1).font = Font(name='Calibri', size=11, bold=True)
            ws_duration.cell(row=row, column=3).value = total_events
            ws_duration.cell(row=row, column=3).font = Font(name='Calibri', size=11, bold=True)
            ws_duration.cell(row=row, column=4).value = "100.0%"
            ws_duration.cell(row=row, column=4).font = Font(name='Calibri', size=11, bold=True)
            
            for col in range(1, 6):
                ws_duration.cell(row=row, column=col).border = thin_border
                ws_duration.cell(row=row, column=col).alignment = centered_align
            
            # Crear gráfico de barras para categorías de duración
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Distribución de Eventos por Duración"
            chart.y_axis.title = "Número de eventos"
            chart.x_axis.title = "Categorías"
            
            data = Reference(ws_duration, min_col=3, min_row=3, max_row=7, max_col=3)
            cats = Reference(ws_duration, min_col=1, min_row=4, max_row=7)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.shape = 4
            
            # Personalizar colores de las barras (aproximación - los colores exactos no son configurables directamente)
            s = chart.series[0]
            s.graphicalProperties.solidFill = "4472C4"  # Azul
            
            ws_duration.add_chart(chart, "A10")
            
            # --- DURACIÓN PROMEDIO POR UBICACIÓN ---
            row_offset = 25  # Ajustar según tamaño del gráfico anterior
            
            ws_duration.merge_cells(f'A{row_offset}:F{row_offset}')
            ws_duration[f'A{row_offset}'] = "DURACIÓN PROMEDIO POR UBICACIÓN"
            ws_duration[f'A{row_offset}'].font = subtitle_font
            ws_duration[f'A{row_offset}'].alignment = centered_align
            
            # Encabezados para duración promedio
            avg_headers = ["Ubicación", "Duración Promedio", "Categoría"]
            for col, header in enumerate(avg_headers, 1):
                ws_duration.cell(row=row_offset+2, column=col).value = header
                ws_duration.cell(row=row_offset+2, column=col).font = header_font
                ws_duration.cell(row=row_offset+2, column=col).fill = header_fill
                ws_duration.cell(row=row_offset+2, column=col).alignment = centered_align
                ws_duration.cell(row=row_offset+2, column=col).border = thin_border
            
            # Ajustar ancho para esta sección
            ws_duration.column_dimensions['A'].width = 30
            ws_duration.column_dimensions['B'].width = 20
            ws_duration.column_dimensions['C'].width = 15
            
            # Datos de duración promedio por ubicación
            if 'duracion_promedio' in stats and len(stats['duracion_promedio']) > 0:
                # Ordenar por duración promedio (descendente)
                sorted_data = sorted(stats['duracion_promedio'], key=lambda x: x['duracion_promedio'], reverse=True)
                
                for i, item in enumerate(sorted_data):
                    row = row_offset + 3 + i
                    location = item['nombre_posicion']
                    duration = item['duracion_promedio']
                    
                    # Determinar categoría
                    if duration < 5:
                        category = "Bajo"
                        category_fill = PatternFill("solid", fgColor="E8F5E9")  # Verde claro
                        category_font = Font(name='Calibri', size=11, color="2E7D32", bold=True)  # Verde oscuro
                    elif duration < 20:
                        category = "Moderado"
                        category_fill = PatternFill("solid", fgColor="FFF8E1")  # Amarillo claro
                        category_font = Font(name='Calibri', size=11, color="F57F17", bold=True)  # Naranja
                    elif duration < 60:
                        category = "Alto"
                        category_fill = PatternFill("solid", fgColor="FFF3E0")  # Naranja claro
                        category_font = Font(name='Calibri', size=11, color="E65100", bold=True)  # Naranja oscuro
                    else:
                        category = "Crítico"
                        category_fill = PatternFill("solid", fgColor="FFEBEE")  # Rojo claro
                        category_font = Font(name='Calibri', size=11, color="B71C1C", bold=True)  # Rojo oscuro
                    
                    # Ubicación
                    cell = ws_duration.cell(row=row, column=1)
                    cell.value = location
                    cell.font = normal_font
                    cell.alignment = left_align
                    cell.border = thin_border
                    
                    # Duración
                    cell = ws_duration.cell(row=row, column=2)
                    cell.value = f"{duration:.1f} segundos"
                    cell.font = normal_font
                    cell.alignment = centered_align
                    cell.border = thin_border
                    
                    # Categoría
                    cell = ws_duration.cell(row=row, column=3)
                    cell.value = category
                    cell.font = category_font
                    cell.fill = category_fill
                    cell.alignment = centered_align
                    cell.border = thin_border
                    
                    # Alternar color de fondo para filas
                    if i % 2 == 0:
                        for col in range(1, 3):
                            ws_duration.cell(row=row, column=col).fill = PatternFill("solid", fgColor="F7FAFC")
            
            # --- HOJA: EVENTOS DE LARGA DURACIÓN ---
            if 'eventos_largos' in stats and len(stats['eventos_largos']) > 0:
                ws_long = wb.create_sheet("Eventos Largos")
                
                # Título
                ws_long.merge_cells('A1:F1')
                ws_long['A1'] = "EVENTOS DE LARGA DURACIÓN"
                ws_long['A1'].font = title_font
                ws_long['A1'].fill = title_fill
                ws_long['A1'].alignment = centered_align
                
                # Texto explicativo
                ws_long.merge_cells('A2:F2')
                ws_long['A2'] = "Eventos con duración superior a 20 segundos, ordenados de mayor a menor duración"
                ws_long['A2'].font = normal_font
                ws_long['A2'].alignment = left_align
                
                # Encabezados para eventos largos
                long_headers = ["Ubicación", "Fecha", "Hora", "Duración", "Estado"]
                for col, header in enumerate(long_headers, 1):
                    ws_long.cell(row=4, column=col).value = header
                    ws_long.cell(row=4, column=col).font = header_font
                    ws_long.cell(row=4, column=col).fill = header_fill
                    ws_long.cell(row=4, column=col).alignment = centered_align
                    ws_long.cell(row=4, column=col).border = thin_border
                
                # Ajustar anchos de columna
                ws_long.column_dimensions['A'].width = 30
                ws_long.column_dimensions['B'].width = 12
                ws_long.column_dimensions['C'].width = 12
                ws_long.column_dimensions['D'].width = 12
                ws_long.column_dimensions['E'].width = 15
                
                # Llenar datos de eventos largos
                for i, event in enumerate(stats['eventos_largos']):
                    row = 5 + i
                    
                    # Formatear duración para mejor legibilidad
                    duration = event['duracion_segundos']
                    if duration >= 60:
                        minutes = int(duration // 60)
                        seconds = int(duration % 60)
                        duration_str = f"{minutes}m {seconds}s"
                        duration_fill = PatternFill("solid", fgColor="FFEBEE")  # Rojo claro
                        duration_font = Font(name='Calibri', size=11, color="D32F2F", bold=True)  # Rojo oscuro
                    else:
                        duration_str = f"{int(duration)}s"
                        duration_fill = PatternFill("solid", fgColor="FFF8E1")  # Amarillo claro
                        duration_font = Font(name='Calibri', size=11, color="F57F17", bold=True)  # Naranja oscuro
                    
                    # Estado revisado o pendiente
                    status_str = "✓ Revisado" if event['revisado'] else "⚠ Pendiente"
                    status_font = Font(name='Calibri', size=11, color="006633", bold=True) if event['revisado'] else Font(name='Calibri', size=11, color="CC3300", bold=True)
                    
                    # Ubicación
                    cell = ws_long.cell(row=row, column=1)
                    cell.value = event['nombre_posicion']
                    cell.font = normal_font
                    cell.alignment = left_align
                    cell.border = thin_border
                    
                    # Fecha
                    cell = ws_long.cell(row=row, column=2)
                    cell.value = event['fecha_evento']
                    cell.font = normal_font
                    cell.alignment = centered_align
                    cell.border = thin_border
                    
                    # Hora
                    cell = ws_long.cell(row=row, column=3)
                    cell.value = event['hora_evento']
                    cell.font = normal_font
                    cell.alignment = centered_align
                    cell.border = thin_border
                    
                    # Duración
                    cell = ws_long.cell(row=row, column=4)
                    cell.value = duration_str
                    cell.font = duration_font
                    cell.fill = duration_fill
                    cell.alignment = centered_align
                    cell.border = thin_border
                    
                    # Estado
                    cell = ws_long.cell(row=row, column=5)
                    cell.value = status_str
                    cell.font = status_font
                    cell.alignment = centered_align
                    cell.border = thin_border
                    
                    # Alternar colores de fila
                    if i % 2 == 0:
                        for col in [1, 2, 3, 5]:  # Excluimos la columna de duración que ya tiene su propio color
                            ws_long.cell(row=row, column=col).fill = PatternFill("solid", fgColor="F0F5FA")

    # --- IMPLEMENTACIÓN ESPECÍFICA PARA ANÁLISIS DE CÁMARAS ---
    if report_type == "cameras-analysis":
        # Variables para mantener la posición actual
        current_row = 11
        
        # Título de sección
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = "ANÁLISIS DE ACTIVIDAD POR CÁMARA"
        ws[f'A{current_row}'].font = title_font
        ws[f'A{current_row}'].fill = header_fill
        ws[f'A{current_row}'].alignment = centered_align
        current_row += 2
        
        # 1. Crear hoja de resumen de cámaras
        ws_cameras = wb.create_sheet("Actividad por Cámara")
        
        # Título en la hoja de cámaras
        ws_cameras.merge_cells('A1:F1')
        ws_cameras['A1'] = "ACTIVIDAD REGISTRADA POR CÁMARA"
        ws_cameras['A1'].font = title_font
        ws_cameras['A1'].fill = title_fill
        ws_cameras['A1'].alignment = centered_align
        
        # Encabezados para resumen de cámaras
        cameras_headers = ["Cámara", "Total Alertas", "% del Total", "Nivel de Actividad", "Última Actividad"]
        for col, header in enumerate(cameras_headers, 1):
            ws_cameras.cell(row=3, column=col).value = header
            ws_cameras.cell(row=3, column=col).font = header_font
            ws_cameras.cell(row=3, column=col).fill = header_fill
            ws_cameras.cell(row=3, column=col).alignment = centered_align
            ws_cameras.cell(row=3, column=col).border = thin_border
        
        # Ajustar anchos de columna
        ws_cameras.column_dimensions['A'].width = 30
        ws_cameras.column_dimensions['B'].width = 15
        ws_cameras.column_dimensions['C'].width = 12
        ws_cameras.column_dimensions['D'].width = 20
        ws_cameras.column_dimensions['E'].width = 20
        
        # Datos de cámaras (simulados o reales si están disponibles)
        if charts_data and 'by_camera' in charts_data and len(charts_data['by_camera']['labels']) > 0:
            # Usar datos reales de charts_data
            labels = charts_data['by_camera']['labels']
            data = charts_data['by_camera']['data']
            total = sum(data)
            
            # Niveles de actividad
            activity_levels = []
            for value in data:
                percentage = (value/total*100) if total > 0 else 0
                if percentage > 40:
                    activity_levels.append(("ALTA", PatternFill("solid", fgColor="FFCDD2"), Font(name='Calibri', size=11, color="B71C1C", bold=True)))
                elif percentage > 20:
                    activity_levels.append(("MEDIA", PatternFill("solid", fgColor="FFF9C4"), Font(name='Calibri', size=11, color="F57F17", bold=True)))
                else:
                    activity_levels.append(("BAJA", PatternFill("solid", fgColor="C8E6C9"), Font(name='Calibri', size=11, color="2E7D32", bold=True)))
            
            # Última actividad (simulada, debería venir en los datos reales)
            last_activities = []
            from datetime import datetime, timedelta
            for i in range(len(labels)):
                days_ago = i % 5  # Simular entre 0 y 4 días atrás
                last_activities.append((datetime.now() - timedelta(days=days_ago)).strftime("%d/%m/%Y %H:%M"))
            
            # Llenar datos en la hoja
            for i, (label, count) in enumerate(zip(labels, data)):
                row = 4 + i
                percentage = (count/total*100) if total > 0 else 0
                
                # Cámara
                cell = ws_cameras.cell(row=row, column=1)
                cell.value = label
                cell.font = normal_font
                cell.alignment = left_align
                cell.border = thin_border
                
                # Total Alertas
                cell = ws_cameras.cell(row=row, column=2)
                cell.value = count
                cell.font = normal_font
                cell.alignment = centered_align
                cell.border = thin_border
                
                # Porcentaje
                cell = ws_cameras.cell(row=row, column=3)
                cell.value = f"{percentage:.1f}%"
                cell.font = normal_font
                cell.alignment = centered_align
                cell.border = thin_border
                
                # Nivel de Actividad
                cell = ws_cameras.cell(row=row, column=4)
                cell.value = activity_levels[i][0]
                cell.font = activity_levels[i][2]
                cell.fill = activity_levels[i][1]
                cell.alignment = centered_align
                cell.border = thin_border
                
                # Última Actividad
                cell = ws_cameras.cell(row=row, column=5)
                cell.value = last_activities[i]
                cell.font = normal_font
                cell.alignment = centered_align
                cell.border = thin_border
                
                # Alternar colores de fondo para filas
                if i % 2 == 0:
                    for col in [1, 2, 3, 5]:  # Excluir columna de nivel que ya tiene color
                        ws_cameras.cell(row=row, column=col).fill = PatternFill("solid", fgColor="F7FAFC")
            
            # Crear gráfico de barras para actividad de cámaras
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Distribución de Alertas por Cámara"
            chart.y_axis.title = "Número de Alertas"
            chart.x_axis.title = "Cámaras"
            
            data = Reference(ws_cameras, min_col=2, min_row=3, max_row=3+len(labels), max_col=2)
            cats = Reference(ws_cameras, min_col=1, min_row=4, max_row=3+len(labels))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.shape = 4
            
            # Personalizar colores de las barras
            s = chart.series[0]
            s.graphicalProperties.solidFill = "4472C4"  # Azul
            
            ws_cameras.add_chart(chart, "A" + str(6 + len(labels)))
            
            # Referencia en la hoja principal
            ws[f'A{current_row}'] = "Distribución de alertas por cámara"
            apply_subheader_style(ws[f'A{current_row}'])
            current_row += 1
            
            ws[f'A{current_row}'] = f"Se registraron un total de {total} alertas en {len(labels)} cámaras."
            apply_normal_style(ws[f'A{current_row}'])
            current_row += 1
            
            ws[f'A{current_row}'] = "Ver detalles completos en la hoja 'Actividad por Cámara'"
            apply_normal_style(ws[f'A{current_row}'])
            current_row += 2
        else:
            # Si no hay datos, mostrar mensaje
            ws[f'A{current_row}'] = "No hay datos de actividad por cámara disponibles para el período seleccionado."
            apply_normal_style(ws[f'A{current_row}'])
            current_row += 2
        
        # 2. Crear hoja de análisis de patrones horarios
        ws_patterns = wb.create_sheet("Patrones Horarios")
        
        # Título en la hoja de patrones
        ws_patterns.merge_cells('A1:F1')
        ws_patterns['A1'] = "PATRONES HORARIOS DE ACTIVIDAD"
        ws_patterns['A1'].font = title_font
        ws_patterns['A1'].fill = title_fill
        ws_patterns['A1'].alignment = centered_align
        
        # Si hay datos de patrones horarios
        if charts_data and 'hourly_pattern' in charts_data and len(charts_data['hourly_pattern'].get('labels', [])) > 0:
            # Usar datos reales
            hour_labels = charts_data['hourly_pattern']['labels']
            hour_data = charts_data['hourly_pattern']['data']
            
            # Encabezados para patrones horarios
            ws_patterns['A3'] = "Hora"
            ws_patterns['B3'] = "Número de Alertas"
            ws_patterns['C3'] = "% del Total"
            
            # Aplicar estilo a encabezados
            apply_header_style(ws_patterns['A3'])
            apply_header_style(ws_patterns['B3'])
            apply_header_style(ws_patterns['C3'])
            
            # Ajustar anchos de columna
            ws_patterns.column_dimensions['A'].width = 15
            ws_patterns.column_dimensions['B'].width = 20
            ws_patterns.column_dimensions['C'].width = 15
            
            # Calcular total
            total_hourly = sum(hour_data)
            
            # Llenar datos
            for i, (hour, count) in enumerate(zip(hour_labels, hour_data)):
                row = 4 + i
                percentage = (count/total_hourly*100) if total_hourly > 0 else 0
                
                # Hora
                ws_patterns[f'A{row}'] = hour
                ws_patterns[f'A{row}'].alignment = centered_align
                
                # Número de Alertas
                ws_patterns[f'B{row}'] = count
                ws_patterns[f'B{row}'].alignment = centered_align
                
                # Porcentaje
                ws_patterns[f'C{row}'] = f"{percentage:.1f}%"
                ws_patterns[f'C{row}'].alignment = centered_align
                
                # Alternar colores de fondo
                if i % 2 == 0:
                    for col in ['A', 'B', 'C']:
                        ws_patterns[f'{col}{row}'].fill = PatternFill("solid", fgColor="F7FAFC")
                
                # Aplicar bordes
                for col in ['A', 'B', 'C']:
                    ws_patterns[f'{col}{row}'].border = thin_border
            
            # Crear gráfico de línea para patrones horarios
            line = LineChart()
            line.title = "Distribución Horaria de Alertas"
            line.style = 12
            line.y_axis.title = "Número de Alertas"
            line.x_axis.title = "Hora del Día"
            
            data = Reference(ws_patterns, min_col=2, min_row=3, max_row=3+len(hour_labels), max_col=2)
            cats = Reference(ws_patterns, min_col=1, min_row=4, max_row=3+len(hour_labels))
            line.add_data(data, titles_from_data=True)
            line.set_categories(cats)
            
            # Personalizar línea
            s = line.series[0]
            s.graphicalProperties.solidFill = "4472C4"  # Azul
            
            ws_patterns.add_chart(line, "E3")
            
            # Referencia en la hoja principal
            ws[f'A{current_row}'] = "Patrones horarios de actividad"
            apply_subheader_style(ws[f'A{current_row}'])
            current_row += 1
            
            # Encontrar hora pico
            max_index = hour_data.index(max(hour_data))
            peak_hour = hour_labels[max_index]
            peak_percentage = (max(hour_data)/total_hourly*100) if total_hourly > 0 else 0
            
            ws[f'A{current_row}'] = f"Hora pico de actividad: {peak_hour} ({peak_percentage:.1f}% de las alertas)"
            apply_normal_style(ws[f'A{current_row}'])
            current_row += 1
            
            ws[f'A{current_row}'] = "Ver análisis completo en la hoja 'Patrones Horarios'"
            apply_normal_style(ws[f'A{current_row}'])
            current_row += 2
        else:
            # Si no hay datos, mostrar mensaje
            ws_patterns['A3'] = "No hay datos de patrones horarios disponibles para el período seleccionado."
            ws_patterns['A3'].font = normal_font
            
            ws[f'A{current_row}'] = "No hay datos de patrones horarios disponibles para el período seleccionado."
            apply_normal_style(ws[f'A{current_row}'])
            current_row += 2
        
        # 3. Listado de alertas recientes
        if alerts_data and len(alerts_data) > 0:
            # Título de sección
            ws.merge_cells(f'A{current_row}:H{current_row}')
            ws[f'A{current_row}'] = "ALERTAS RECIENTES"
            ws[f'A{current_row}'].font = title_font
            ws[f'A{current_row}'].fill = header_fill
            ws[f'A{current_row}'].alignment = centered_align
            current_row += 2
            
            # Mostrar las 5 alertas más recientes en la hoja principal
            max_alerts = min(5, len(alerts_data))
            
            # Encabezados
            headers = ["Fecha", "Hora", "Cámara", "Evento"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                apply_header_style(cell)
            current_row += 1
            
            # Mostrar alertas recientes
            for i in range(max_alerts):
                alert = alerts_data[i]
                
                # Fecha (asumiendo que hay campo fecha_evento)
                cell = ws.cell(row=current_row, column=1)
                cell.value = alert.get('fecha_evento', 'N/D')
                cell.font = normal_font
                cell.border = thin_border
                
                # Hora (asumiendo que hay campo hora_evento)
                cell = ws.cell(row=current_row, column=2)
                cell.value = alert.get('hora_evento', 'N/D')
                cell.font = normal_font
                cell.border = thin_border
                
                # Cámara (asumiendo que hay campo nombre_posicion o ubicacion)
                cell = ws.cell(row=current_row, column=3)
                cell.value = alert.get('nombre_posicion', alert.get('ubicacion', 'N/D'))
                cell.font = normal_font
                cell.border = thin_border
                
                # Evento (asumiendo que hay campo tipo_evento o descripcion)
                cell = ws.cell(row=current_row, column=4)
                cell.value = alert.get('tipo_evento', alert.get('descripcion', 'Alerta'))
                cell.font = normal_font
                cell.border = thin_border
                
                # Alternar colores de fila
                if i % 2 == 0:
                    for col in range(1, 5):
                        ws.cell(row=current_row, column=col).fill = PatternFill("solid", fgColor="F7FAFC")
                
                current_row += 1
            
            # Crear una hoja dedicada para todas las alertas
            ws_alerts = wb.create_sheet("Alertas Detalladas")
            
            # Título
            ws_alerts.merge_cells('A1:F1')
            ws_alerts['A1'] = "LISTADO COMPLETO DE ALERTAS"
            ws_alerts['A1'].font = title_font
            ws_alerts['A1'].fill = title_fill
            ws_alerts['A1'].alignment = centered_align
            
            # Obtener todas las claves posibles para los encabezados
            all_keys = set()
            for alert in alerts_data:
                all_keys.update(alert.keys())
            
            # Lista ordenada de encabezados prioritarios
            priority_headers = [
                'fecha_evento', 'hora_evento', 'nombre_posicion', 'ubicacion', 
                'tipo_evento', 'descripcion', 'duracion_segundos', 'nivel_alerta'
            ]
            
            # Ordenar headers: primero los prioritarios, luego el resto alfabéticamente
            headers = [h for h in priority_headers if h in all_keys]
            remaining_headers = sorted(list(all_keys - set(headers)))
            headers.extend(remaining_headers)
            
            # Encabezados en la hoja de alertas
            for col, header in enumerate(headers, 1):
                # Convertir nombre técnico a nombre más legible
                display_name = header.replace('_', ' ').title()
                
                cell = ws_alerts.cell(row=3, column=col)
                cell.value = display_name
                apply_header_style(cell)
                
                # Ajustar ancho de columna
                col_letter = get_column_letter(col)
                ws_alerts.column_dimensions[col_letter].width = max(15, len(display_name) + 2)
            
            # Llenar datos de alertas
            for i, alert in enumerate(alerts_data):
                row = 4 + i
                
                for j, header in enumerate(headers, 1):
                    cell = ws_alerts.cell(row=row, column=j)
                    cell.value = alert.get(header, '')
                    cell.font = normal_font
                    cell.border = thin_border
                    
                    # Si es duración, aplicar formato especial
                    if header == 'duracion_segundos' and alert.get(header):
                        duration = float(alert.get(header))
                        if duration >= 60:
                            cell.fill = PatternFill("solid", fgColor="FFCDD2")  # Rojo claro
                        elif duration >= 20:
                            cell.fill = PatternFill("solid", fgColor="FFF9C4")  # Amarillo claro
                    
                    # Si es nivel de alerta, aplicar formato por nivel
                    if header == 'nivel_alerta' and alert.get(header):
                        nivel = str(alert.get(header)).upper()
                        if nivel == 'ALTA' or nivel == 'ALTO':
                            cell.fill = PatternFill("solid", fgColor="FFCDD2")  # Rojo claro
                            cell.font = Font(name='Calibri', size=11, color="B71C1C", bold=True)
                        elif nivel == 'MEDIA' or nivel == 'MEDIO':
                            cell.fill = PatternFill("solid", fgColor="FFF9C4")  # Amarillo claro
                            cell.font = Font(name='Calibri', size=11, color="F57F17", bold=True)
                
                # Alternar colores de fila
                if i % 2 == 0:
                    for j, _ in enumerate(headers, 1):
                        # No sobrescribir celdas que ya tienen formato especial
                        cell = ws_alerts.cell(row=row, column=j)
                        if not cell.fill.fgColor.rgb:
                            cell.fill = PatternFill("solid", fgColor="F7FAFC")
            
            # Referencia en la hoja principal
            current_row += 1
            ws[f'A{current_row}'] = f"Ver listado completo de {len(alerts_data)} alertas en la hoja 'Alertas Detalladas'"
            apply_normal_style(ws[f'A{current_row}'])
        else:
            # Si no hay alertas
            ws[f'A{current_row}'] = "No hay alertas registradas para el período seleccionado."
            apply_normal_style(ws[f'A{current_row}'])

    # --- IMPLEMENTACIÓN DEL REPORTE COMPLETO INTEGRADO ---
    if report_type == "complete-integrated":
        # --- HOJA PRINCIPAL: RESUMEN EJECUTIVO ---
        # Crear una única hoja principal con el resumen ejecutivo
        current_row = 11
        
        # Título de sección
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = "RESUMEN EJECUTIVO"
        ws[f'A{current_row}'].font = title_font
        ws[f'A{current_row}'].fill = title_fill
        ws[f'A{current_row}'].alignment = centered_align
        current_row += 2
        
        # --- SECCIÓN 1: KPIs PRINCIPALES ---
        ws[f'A{current_row}'] = "INDICADORES CLAVE DE RENDIMIENTO"
        ws[f'A{current_row}'].font = subtitle_font
        current_row += 1
        
        # Crear tabla de 2x3 para los KPIs principales
        kpi_headers = ["Indicador", "Valor"]
        for col, header in enumerate(kpi_headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            apply_header_style(cell)
        current_row += 1
        
        # Calcular KPIs
        total_events = 0
        total_cameras = 0
        avg_duration = 0
        critical_events = 0
        
        # Total de eventos
        if stats and 'duracion_categorias' in stats:
            total_events = sum(stats['duracion_categorias'][cat] for cat in ['bajo', 'moderado', 'alto', 'critico'])
            critical_events = stats['duracion_categorias']['alto'] + stats['duracion_categorias']['critico']
        
        # Total de cámaras
        if charts_data and 'by_camera' in charts_data:
            total_cameras = len(charts_data['by_camera']['labels'])
        
        # Duración promedio
        if stats and 'duracion_promedio_general' in stats:
            avg_duration = stats['duracion_promedio_general']
        elif stats and 'duracion_promedio' in stats and len(stats['duracion_promedio']) > 0:
            durations = [item['duracion_promedio'] for item in stats['duracion_promedio']]
            avg_duration = sum(durations) / len(durations) if durations else 0
        
        # Llenar KPIs
        kpi_data = [
            ["Total de Eventos", total_events],
            ["Eventos Críticos", critical_events],
            ["Duración Promedio", f"{avg_duration:.1f} segundos" if avg_duration else "N/D"],
            ["Total de Cámaras", total_cameras]
        ]
        
        for i, (indicator, value) in enumerate(kpi_data):
            # Indicador
            cell = ws.cell(row=current_row, column=1)
            cell.value = indicator
            cell.font = normal_font
            cell.border = thin_border
            
            # Valor
            cell = ws.cell(row=current_row, column=2)
            cell.value = value
            cell.font = Font(name='Calibri', size=11, bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            
            # Alternar colores
            if i % 2 == 0:
                for col in range(1, 3):
                    ws.cell(row=current_row, column=col).fill = PatternFill("solid", fgColor="F7FAFC")
            
            current_row += 1
        
        current_row += 2
        
    # --- SECCIÓN 2: GRÁFICO UNIFICADO DE DURACIÓN Y CÁMARAS ---
    ws_overview = wb.create_sheet("Vista General")
    
    # Título en la hoja de resumen
    ws_overview.merge_cells('A1:F1')
    ws_overview['A1'] = "VISTA GENERAL INTEGRADA"
    ws_overview['A1'].font = title_font
    ws_overview['A1'].fill = title_fill
    ws_overview['A1'].alignment = centered_align
    
    # 2.1 Gráfico de Duración
    ws_overview.merge_cells('A3:C3')
    ws_overview['A3'] = "DISTRIBUCIÓN DE EVENTOS POR DURACIÓN"
    ws_overview['A3'].font = subtitle_font
    ws_overview['A3'].alignment = centered_align
    
    # Encabezados para categorías de duración
    duration_headers = ["Categoría", "Eventos", "Porcentaje"]
    for col, header in enumerate(duration_headers, 1):
        ws_overview.cell(row=4, column=col).value = header
        ws_overview.cell(row=4, column=col).font = header_font
        ws_overview.cell(row=4, column=col).fill = header_fill
        ws_overview.cell(row=4, column=col).alignment = centered_align
        ws_overview.cell(row=4, column=col).border = thin_border
    
    # Datos de duración
    if stats and 'duracion_categorias' in stats:
        duration_data = [
            ["Bajo (< 5s)", stats['duracion_categorias']['bajo']],
            ["Moderado (5-20s)", stats['duracion_categorias']['moderado']],
            ["Alto (20-60s)", stats['duracion_categorias']['alto']],
            ["Crítico (> 60s)", stats['duracion_categorias']['critico']],
        ]
        
        # Calcular porcentajes
        total_events = sum(stats['duracion_categorias'][cat] for cat in ['bajo', 'moderado', 'alto', 'critico'])
        
        # Colores para las categorías
        category_fills = {
            "Bajo (< 5s)": PatternFill("solid", fgColor="66BB6A"),
            "Moderado (5-20s)": PatternFill("solid", fgColor="FFC107"),
            "Alto (20-60s)": PatternFill("solid", fgColor="FF9800"),
            "Crítico (> 60s)": PatternFill("solid", fgColor="F44336")
        }
        
        # Rellenar datos en la hoja
        for i, row_data in enumerate(duration_data):
            category = row_data[0]
            count = row_data[1]
            percentage = (count/total_events*100) if total_events > 0 else 0
            
            row = 5 + i
            # Categoría
            cell = ws_overview.cell(row=row, column=1)
            cell.value = category
            cell.font = normal_font
            cell.border = thin_border
            cell.fill = category_fills[category]
            
            # Eventos
            cell = ws_overview.cell(row=row, column=2)
            cell.value = count
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            
            # Porcentaje
            cell = ws_overview.cell(row=row, column=3)
            cell.value = f"{percentage:.1f}%"
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
            pie = PieChart()
            pie.title = "Distribución por Duración"
            labels = Reference(ws_overview, min_col=1, min_row=5, max_row=8)
            data = Reference(ws_overview, min_col=2, min_row=4, max_row=8)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)

# Establecer un tamaño específico para el gráfico de pastel
            pie.width = 15  # Unidades en centímetros
            pie.height = 10

# Añadir gráfico en posición A10
            ws_overview.add_chart(pie, "A10")
        
        # Referencia en la hoja principal
        ws[f'A{current_row}'] = "Distribución de Eventos por Duración"
        apply_subheader_style(ws[f'A{current_row}'])
        current_row += 1
        
        # Mensaje resumen en hoja principal
        critical_percentage = ((stats['duracion_categorias']['alto'] + stats['duracion_categorias']['critico'])/total_events*100) if total_events > 0 else 0
        ws[f'A{current_row}'] = f"El {critical_percentage:.1f}% de los eventos requieren atención (duración > 20s)."
        apply_normal_style(ws[f'A{current_row}'])
        current_row += 2
    
    # 2.2 Gráfico de Cámaras (en la misma hoja)
    ws_overview.merge_cells('D3:F3')
    ws_overview['D3'] = "ACTIVIDAD POR CÁMARA (TOP 5)"
    ws_overview['D3'].font = subtitle_font
    ws_overview['D3'].alignment = centered_align
    
    # Encabezados para cámaras
    camera_headers = ["Cámara", "Alertas", "% del Total"]
    for col, header in enumerate(camera_headers, 4):  # Continuar desde columna D
        ws_overview.cell(row=4, column=col).value = header
        ws_overview.cell(row=4, column=col).font = header_font
        ws_overview.cell(row=4, column=col).fill = header_fill
        ws_overview.cell(row=4, column=col).alignment = centered_align
        ws_overview.cell(row=4, column=col).border = thin_border
    
    # Datos de cámaras
    if charts_data and 'by_camera' in charts_data and len(charts_data['by_camera']['labels']) > 0:
        # Usar datos reales de charts_data
        labels = charts_data['by_camera']['labels']
        data = charts_data['by_camera']['data']
        total = sum(data)
        
        # Obtener top 5 cámaras
        camera_data = list(zip(labels, data))
        camera_data.sort(key=lambda x: x[1], reverse=True)
        top_cameras = camera_data[:5]
        
        # Llenar datos en la hoja
        for i, (label, count) in enumerate(top_cameras):
            row = 5 + i
            percentage = (count/total*100) if total > 0 else 0
            
            # Cámara
            cell = ws_overview.cell(row=row, column=4)
            cell.value = label
            cell.font = normal_font
            cell.border = thin_border
            
            # Alertas
            cell = ws_overview.cell(row=row, column=5)
            cell.value = count
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            
            # Porcentaje
            cell = ws_overview.cell(row=row, column=6)
            cell.value = f"{percentage:.1f}%"
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            
            # Alternar colores
            if i % 2 == 0:
                for col in range(4, 7):
                    ws_overview.cell(row=row, column=col).fill = PatternFill("solid", fgColor="F7FAFC")
        
        # Crear gráfico de barras para top 5 cámaras
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Top 5 Cámaras por Actividad"
        chart.y_axis.title = "Alertas"
        
        data = Reference(ws_overview, min_col=5, min_row=4, max_row=9, max_col=5)
        cats = Reference(ws_overview, min_col=4, min_row=5, max_row=9)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        # Personalizar colores de las barras
        s = chart.series[0]
        s.graphicalProperties.solidFill = "4472C4"  # Azul
        
        # Colocar el gráfico de barras en la parte derecha
        ws_overview.add_chart(chart, "A30")
        
        # Mensaje resumen en hoja principal
        top_camera = top_cameras[0][0]
        top_percentage = (top_cameras[0][1]/total*100) if total > 0 else 0
        
        ws[f'A{current_row}'] = "Distribución de Actividad por Cámara"
        apply_subheader_style(ws[f'A{current_row}'])
        current_row += 1
        
        ws[f'A{current_row}'] = f"La cámara '{top_camera}' registra la mayor actividad ({top_percentage:.1f}% del total)."
        apply_normal_style(ws[f'A{current_row}'])
        current_row += 2
    
    # --- SECCIÓN 3: PATRONES HORARIOS ---
    ws[f'A{current_row}'] = "PATRONES HORARIOS"
    apply_subheader_style(ws[f'A{current_row}'])
    current_row += 1
    
    if charts_data and 'hourly_pattern' in charts_data and len(charts_data['hourly_pattern'].get('labels', [])) > 0:
        # Crear gráfico en hoja separada
        ws_patterns = wb.create_sheet("Patrones Horarios")
        
        # Título
        ws_patterns.merge_cells('A1:F1')
        ws_patterns['A1'] = "PATRONES HORARIOS DE ACTIVIDAD"
        ws_patterns['A1'].font = title_font
        ws_patterns['A1'].fill = title_fill
        ws_patterns['A1'].alignment = centered_align
        
        # Usar datos reales
        hour_labels = charts_data['hourly_pattern']['labels']
        hour_data = charts_data['hourly_pattern']['data']
        
        # Encabezados
        ws_patterns['A3'] = "Hora"
        ws_patterns['B3'] = "Alertas"
        apply_header_style(ws_patterns['A3'])
        apply_header_style(ws_patterns['B3'])
        
        # Ajustar anchos de columna
        ws_patterns.column_dimensions['A'].width = 15
        ws_patterns.column_dimensions['B'].width = 15
        
        # Calcular total y encontrar hora pico
        total_hourly = sum(hour_data)
        max_index = hour_data.index(max(hour_data)) if hour_data else 0
        peak_hour = hour_labels[max_index] if hour_labels else "N/D"
        peak_count = max(hour_data) if hour_data else 0
        
        # Llenar datos
        for i, (hour, count) in enumerate(zip(hour_labels, hour_data)):
            row = 4 + i
            
            # Hora
            ws_patterns[f'A{row}'] = hour
            ws_patterns[f'A{row}'].alignment = centered_align
            ws_patterns[f'A{row}'].border = thin_border
            
            # Alertas
            ws_patterns[f'B{row}'] = count
            ws_patterns[f'B{row}'].alignment = centered_align
            ws_patterns[f'B{row}'].border = thin_border
            
            # Destacar hora pico
            if i == max_index:
                ws_patterns[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
                ws_patterns[f'B{row}'].font = Font(name='Calibri', size=11, bold=True)
                ws_patterns[f'A{row}'].fill = PatternFill("solid", fgColor="E3F2FD")  # Azul claro
                ws_patterns[f'B{row}'].fill = PatternFill("solid", fgColor="E3F2FD")  # Azul claro
            else:
                # Alternar colores de fondo
                if i % 2 == 0:
                    ws_patterns[f'A{row}'].fill = PatternFill("solid", fgColor="F7FAFC")
                    ws_patterns[f'B{row}'].fill = PatternFill("solid", fgColor="F7FAFC")
        
        # Crear gráfico de línea
        line = LineChart()
        line.title = "Distribución Horaria de Alertas"
        line.style = 12
        line.y_axis.title = "Alertas"
        line.x_axis.title = "Hora"
        
        data = Reference(ws_patterns, min_col=2, min_row=3, max_row=3+len(hour_labels), max_col=2)
        cats = Reference(ws_patterns, min_col=1, min_row=4, max_row=3+len(hour_labels))
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        
        ws_patterns.add_chart(line, "D3")
        
        # Referencia en la hoja principal
        peak_percentage = (peak_count/total_hourly*100) if total_hourly > 0 else 0
        ws[f'A{current_row}'] = f"Hora pico: {peak_hour} con {peak_count} alertas ({peak_percentage:.1f}% del total)"
        apply_normal_style(ws[f'A{current_row}'])
        current_row += 1
        
        ws[f'A{current_row}'] = "Ver detalles en la hoja 'Patrones Horarios'"
        apply_normal_style(ws[f'A{current_row}'])
        current_row += 2
    else:
        ws[f'A{current_row}'] = "No hay datos de patrones horarios disponibles."
        apply_normal_style(ws[f'A{current_row}'])
        current_row += 2
        
        # --- SECCIÓN 5: ANÁLISIS DE UBICACIONES ---
        if stats and 'duracion_promedio' in stats and len(stats['duracion_promedio']) > 0:
            ws[f'A{current_row}'] = "ANÁLISIS DE UBICACIONES"
            apply_subheader_style(ws[f'A{current_row}'])
            current_row += 1
            
            # Crear hoja de análisis de ubicaciones
            ws_locations = wb.create_sheet("Análisis por Ubicación")
            
            # Título
            ws_locations.merge_cells('A1:F1')
            ws_locations['A1'] = "ANÁLISIS POR UBICACIÓN"
            ws_locations['A1'].font = title_font
            ws_locations['A1'].fill = title_fill
            ws_locations['A1'].alignment = centered_align
            
            # Encabezados
            locations_headers = ["Ubicación", "Total Eventos", "Duración Promedio", "Categoría"]
            for col, header in enumerate(locations_headers, 1):
                ws_locations.cell(row=3, column=col).value = header
                ws_locations.cell(row=3, column=col).font = header_font
                ws_locations.cell(row=3, column=col).fill = header_fill
                ws_locations.cell(row=3, column=col).alignment = centered_align
                ws_locations.cell(row=3, column=col).border = thin_border
            
            # Ajustar anchos
            ws_locations.column_dimensions['A'].width = 30
            ws_locations.column_dimensions['B'].width = 15
            ws_locations.column_dimensions['C'].width = 20
            ws_locations.column_dimensions['D'].width = 15
            
            # Ordenar ubicaciones por duración promedio (descendente)
            locations_data = sorted(stats['duracion_promedio'], key=lambda x: x['duracion_promedio'], reverse=True)
            
            # Llenar datos
            for i, location in enumerate(locations_data):
                row = 4 + i
                
                # Determinar categoría por duración
                duration = location['duracion_promedio']
                if duration < 5:
                    category = "Bajo"
                    category_fill = PatternFill("solid", fgColor="C8E6C9")  # Verde claro
                    category_font = Font(name='Calibri', size=11, color="2E7D32", bold=True)
                elif duration < 20:
                    category = "Moderado"
                    category_fill = PatternFill("solid", fgColor="FFF9C4")  # Amarillo claro
                    category_font = Font(name='Calibri', size=11, color="F57F17", bold=True)
                elif duration < 60:
                    category = "Alto"
                    category_fill = PatternFill("solid", fgColor="FFE0B2")  # Naranja claro
                    category_font = Font(name='Calibri', size=11, color="E65100", bold=True)
                else:
                    category = "Crítico"
                    category_fill = PatternFill("solid", fgColor="FFCDD2")  # Rojo claro
                    category_font = Font(name='Calibri', size=11, color="B71C1C", bold=True)
                
                # Ubicación
                cell = ws_locations.cell(row=row, column=1)
                cell.value = location['nombre_posicion']
                cell.font = normal_font
                cell.border = thin_border
                
                # Total eventos
                cell = ws_locations.cell(row=row, column=2)
                cell.value = location.get('total_eventos', 'N/D')
                cell.value = location.get('total_eventos', 'N/D')
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = centered_align
                
                # Duración promedio
                cell = ws_locations.cell(row=row, column=3)
                cell.value = f"{location['duracion_promedio']:.1f}s"
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = centered_align
                
                # Categoría
                cell = ws_locations.cell(row=row, column=4)
                cell.value = category
                cell.font = category_font
                cell.fill = category_fill
                cell.border = thin_border
                cell.alignment = centered_align
                
                # Alternar colores
                if i % 2 == 0:
                    for col in [1, 2]:  # No alternar columnas con formato especial
                        ws_locations.cell(row=row, column=col).fill = PatternFill("solid", fgColor="F7FAFC")
            
            # Crear gráfico de barras para las ubicaciones más críticas
            # Seleccionar top 10 ubicaciones por duración
            top_locations = locations_data[:10] if len(locations_data) > 10 else locations_data
            
            # Crear gráfico
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Top Ubicaciones por Duración Promedio"
            chart.y_axis.title = "Duración (segundos)"
            
            # Preparar rango de datos para el gráfico
            c1 = ws_locations.cell(row=4, column=1)
            c2 = ws_locations.cell(row=3+len(top_locations), column=1)
            cats = Reference(ws_locations, min_col=1, min_row=4, max_row=3+len(top_locations))
            
            c1 = ws_locations.cell(row=4, column=3)
            c2 = ws_locations.cell(row=3+len(top_locations), column=3)
            data = Reference(ws_locations, min_col=3, min_row=3, max_row=3+len(top_locations))
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            # Personalizar diseño del gráfico
            s = chart.series[0]
            s.graphicalProperties.solidFill = "5A9BD5"  # Azul
            
            # Añadir el gráfico a la hoja
            ws_locations.add_chart(chart, "A20")
            
            # Resumir en la hoja principal
            # Mostrar las 3 ubicaciones más críticas
            top_critical = [loc for loc in locations_data if loc['duracion_promedio'] >= 20][:3]
            
            if top_critical:
                ws[f'A{current_row}'] = "Ubicaciones con duración crítica:"
                apply_normal_style(ws[f'A{current_row}'])
                current_row += 1
                
                for loc in top_critical:
                    duration = loc['duracion_promedio']
                    category = "Alto" if duration < 60 else "Crítico"
                    category_color = "E65100" if duration < 60 else "B71C1C"
                    
                    ws[f'A{current_row}'] = f"• {loc['nombre_posicion']}: {duration:.1f}s ({category})"
                    ws[f'A{current_row}'].font = Font(name='Calibri', size=11, color=category_color, bold=True)
                    current_row += 1
            else:
                ws[f'A{current_row}'] = "No se identificaron ubicaciones con duración crítica."
                apply_normal_style(ws[f'A{current_row}'])
                current_row += 1
            
            ws[f'A{current_row}'] = "Ver análisis completo en la hoja 'Análisis por Ubicación'"
            apply_normal_style(ws[f'A{current_row}'])
            current_row += 2
        else:
            ws[f'A{current_row}'] = "No hay datos de ubicaciones disponibles."
            apply_normal_style(ws[f'A{current_row}'])
            current_row += 2

                # --- SECCIÓN 6: ANÁLISIS DE TIEMPO DE RESPUESTA ---
        if stats and 'tiempo_respuesta' in stats:
            ws[f'A{current_row}'] = "ANÁLISIS DE TIEMPOS DE RESPUESTA"
            apply_subheader_style(ws[f'A{current_row}'])
            current_row += 1
            
            # Crear hoja de análisis de tiempos
            ws_response = wb.create_sheet("Tiempos de Respuesta")
            
            # Título
            ws_response.merge_cells('A1:E1')
            ws_response['A1'] = "ANÁLISIS DE TIEMPOS DE RESPUESTA"
            ws_response['A1'].font = title_font
            ws_response['A1'].fill = title_fill
            ws_response['A1'].alignment = centered_align
            
            # Encabezados
            response_headers = ["Tipo", "Tiempo Promedio", "Tiempo Mínimo", "Tiempo Máximo", "Total Eventos"]
            for col, header in enumerate(response_headers, 1):
                ws_response.cell(row=3, column=col).value = header
                ws_response.cell(row=3, column=col).font = header_font
                ws_response.cell(row=3, column=col).fill = header_fill
                ws_response.cell(row=3, column=col).alignment = centered_align
                ws_response.cell(row=3, column=col).border = thin_border
            
            # Ajustar anchos
            ws_response.column_dimensions['A'].width = 25
            ws_response.column_dimensions['B'].width = 18
            ws_response.column_dimensions['C'].width = 18
            ws_response.column_dimensions['D'].width = 18
            ws_response.column_dimensions['E'].width = 15
            
            # Rellenar datos de tiempos de respuesta
            response_types = stats['tiempo_respuesta'].keys()
            for i, response_type in enumerate(response_types):
                row = 4 + i
                response_data = stats['tiempo_respuesta'][response_type]
                
                # Tipo
                cell = ws_response.cell(row=row, column=1)
                cell.value = response_type.capitalize()
                cell.font = Font(name='Calibri', size=11, bold=True)
                cell.border = thin_border
                
                # Tiempo Promedio
                cell = ws_response.cell(row=row, column=2)
                cell.value = f"{response_data['promedio']:.1f}s"
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = centered_align
                
                # Tiempo Mínimo
                cell = ws_response.cell(row=row, column=3)
                cell.value = f"{response_data['minimo']:.1f}s"
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = centered_align
                
                # Tiempo Máximo
                cell = ws_response.cell(row=row, column=4)
                cell.value = f"{response_data['maximo']:.1f}s"
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = centered_align
                
                # Total Eventos
                cell = ws_response.cell(row=row, column=5)
                cell.value = response_data['total']
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = centered_align
                
                # Alternar colores
                if i % 2 == 0:
                    for col in range(1, 6):
                        ws_response.cell(row=row, column=col).fill = PatternFill("solid", fgColor="F7FAFC")
            
            # Crear gráfico de barras para tiempos de respuesta
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Tiempos de Respuesta por Tipo"
            chart.y_axis.title = "Segundos"
            
            data = Reference(ws_response, min_col=2, min_row=3, max_row=3+len(response_types), max_col=2)
            cats = Reference(ws_response, min_col=1, min_row=4, max_row=3+len(response_types))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            ws_response.add_chart(chart, "A10")
            
            # Mensaje resumen en hoja principal
            avg_times = [(type_name, stats['tiempo_respuesta'][type_name]['promedio']) for type_name in response_types]
            global_avg = sum(time for _, time in avg_times) / len(avg_times) if avg_times else 0
            
            ws[f'A{current_row}'] = f"Tiempo de respuesta promedio global: {global_avg:.1f} segundos"
            apply_normal_style(ws[f'A{current_row}'])
            current_row += 1
            
            # Identificar el tipo más rápido y más lento
            if avg_times:
                fastest = min(avg_times, key=lambda x: x[1])
                slowest = max(avg_times, key=lambda x: x[1])
                
                ws[f'A{current_row}'] = f"Respuesta más rápida: {fastest[0].capitalize()} ({fastest[1]:.1f}s)"
                apply_normal_style(ws[f'A{current_row}'])
                current_row += 1
                
                ws[f'A{current_row}'] = f"Respuesta más lenta: {slowest[0].capitalize()} ({slowest[1]:.1f}s)"
                apply_normal_style(ws[f'A{current_row}'])
                current_row += 1
            
            ws[f'A{current_row}'] = "Ver análisis detallado en la hoja 'Tiempos de Respuesta'"
            apply_normal_style(ws[f'A{current_row}'])
            current_row += 2
        else:
            ws[f'A{current_row}'] = "No hay datos de tiempos de respuesta disponibles."
            apply_normal_style(ws[f'A{current_row}'])
            current_row += 2
        
        # --- SECCIÓN 7: CONCLUSIONES Y RECOMENDACIONES ---
        ws[f'A{current_row}'] = "CONCLUSIONES Y RECOMENDACIONES"
        apply_subheader_style(ws[f'A{current_row}'])
        current_row += 1
        
        # Generar conclusiones basadas en los datos
        conclusions = []
        
        # Conclusión sobre eventos críticos
        if stats and 'duracion_categorias' in stats:
            total_events = sum(stats['duracion_categorias'][cat] for cat in ['bajo', 'moderado', 'alto', 'critico'])
            critical_events = stats['duracion_categorias']['alto'] + stats['duracion_categorias']['critico']
            critical_percentage = (critical_events/total_events*100) if total_events > 0 else 0
            
            if critical_percentage > 30:
                conclusions.append(f"Alto porcentaje de eventos críticos ({critical_percentage:.1f}%). Se recomienda revisar el sistema de respuesta rápida.")
            elif critical_percentage > 15:
                conclusions.append(f"Porcentaje moderado de eventos críticos ({critical_percentage:.1f}%). Se sugiere monitorear las tendencias.")
            else:
                conclusions.append(f"Bajo porcentaje de eventos críticos ({critical_percentage:.1f}%). El sistema está respondiendo adecuadamente.")
        
        # Conclusión sobre cámaras
        if charts_data and 'by_camera' in charts_data and len(charts_data['by_camera']['labels']) > 0:
            labels = charts_data['by_camera']['labels']
            data = charts_data['by_camera']['data']
            total = sum(data)
            
            camera_data = list(zip(labels, data))
            camera_data.sort(key=lambda x: x[1], reverse=True)
            top_camera = camera_data[0][0]
            top_percentage = (camera_data[0][1]/total*100) if total > 0 else 0
            
            if top_percentage > 40:
                conclusions.append(f"La cámara '{top_camera}' concentra una proporción muy alta de alertas ({top_percentage:.1f}%). Se recomienda evaluar esta ubicación.")
            elif top_percentage > 25:
                conclusions.append(f"La cámara '{top_camera}' presenta una concentración significativa de alertas ({top_percentage:.1f}%). Considerar revisión.")
        
        # Conclusión sobre patrones horarios
        if charts_data and 'hourly_pattern' in charts_data and len(charts_data['hourly_pattern'].get('labels', [])) > 0:
            hour_data = charts_data['hourly_pattern']['data']
            hour_labels = charts_data['hourly_pattern']['labels']
            
            max_index = hour_data.index(max(hour_data)) if hour_data else 0
            peak_hour = hour_labels[max_index] if hour_labels else "N/D"
            peak_count = max(hour_data) if hour_data else 0
            total_hourly = sum(hour_data)
            peak_percentage = (peak_count/total_hourly*100) if total_hourly > 0 else 0
            
            if peak_percentage > 20:
                conclusions.append(f"Concentración significativa de alertas a las {peak_hour} ({peak_percentage:.1f}% del total). Se recomienda reforzar la vigilancia en este horario.")
        
        # Conclusión sobre ubicaciones
        if stats and 'duracion_promedio' in stats and len(stats['duracion_promedio']) > 0:
            # Identificar ubicaciones con duración crítica
            critical_locations = [loc for loc in stats['duracion_promedio'] if loc['duracion_promedio'] >= 60]
            high_locations = [loc for loc in stats['duracion_promedio'] if 20 <= loc['duracion_promedio'] < 60]
            
            if critical_locations:
                names = [loc['nombre_posicion'] for loc in critical_locations[:2]]
                names_str = ", ".join(names)
                if len(critical_locations) > 2:
                    names_str += f" y {len(critical_locations)-2} más"
                conclusions.append(f"Las ubicaciones con tiempos críticos ({names_str}) requieren atención inmediata.")
            elif high_locations:
                conclusions.append(f"Se identificaron {len(high_locations)} ubicaciones con tiempos de respuesta altos que deben ser monitoreadas.")
        
        # Generar recomendaciones generales
        recommendations = [
            "Implementar un sistema de alertas tempranas para eventos con duración superior a 20 segundos.",
            "Realizar mantenimiento preventivo en las cámaras con mayor actividad.",
            "Revisar los protocolos de respuesta en las ubicaciones críticas identificadas."
        ]
        
        # Añadir conclusiones al reporte
        if conclusions:
            for i, conclusion in enumerate(conclusions):
                ws[f'A{current_row}'] = f"• {conclusion}"
                apply_normal_style(ws[f'A{current_row}'])
                current_row += 1
        else:
            ws[f'A{current_row}'] = "No hay suficientes datos para generar conclusiones específicas."
            apply_normal_style(ws[f'A{current_row}'])
            current_row += 1
        
        current_row += 1
        ws[f'A{current_row}'] = "Recomendaciones:"
        apply_subheader_style(ws[f'A{current_row}'])
        current_row += 1
        
        # Añadir recomendaciones
        for recommendation in recommendations:
            ws[f'A{current_row}'] = f"• {recommendation}"
            apply_normal_style(ws[f'A{current_row}'])
            current_row += 1
    
    # --- FORMATEO FINAL DEL REPORTE ---
    # Ajustar ancho de columnas en la hoja principal
    for i in range(1, 9):
        col_letter = get_column_letter(i)
        if i == 1:
            ws.column_dimensions[col_letter].width = 35
        else:
            ws.column_dimensions[col_letter].width = 15

    # Guardar el archivo
    wb.save(file_path)
    return file_path

def generate_csv_report(title, user_info, stats, charts_data, alerts_data, report_type, date_range, start_date, end_date, filename):
    import pandas as pd
    import os
    from datetime import datetime
    
    # Crear directorio para reportes si no existe
    reports_dir = os.path.join(app.static_folder, 'reports')
    if not os.path.exists(reports_dir):
        os.makedirs(reports_dir)
    
    # Obtener una ruta de archivo única
    base_path = os.path.join(reports_dir, f"{filename}.csv")
    file_path = get_unique_filename(base_path, 'csv')
    
    # Crear un Excel writer para guardar el reporte
    writer = pd.ExcelWriter(file_path.replace('.csv', '.xlsx'), engine='xlsxwriter')
    
    # Información del reporte
    report_info = {
        'Título del Reporte': [title],
        'Usuario': [f"{user_info['nombres']} {user_info['apellidos']} ({user_info['nombre_usuario']})"],
        'Fecha de Generación': [datetime.now().strftime('%Y-%m-%d')],
        'Hora de Generación': [datetime.now().strftime('%H:%M:%S')],
        'Rango de Fechas': [get_date_range_text(date_range, start_date, end_date)]
    }
    
    report_df = pd.DataFrame(report_info)
    
    # DataFrame principales para guardar en CSV
    dataframes = {}
    
    # 1. DataFrame de información general
    dataframes['Información'] = report_df
    
    # 2. DataFrame de estadísticas de resumen si están disponibles
    if stats and 'total_alerts' in stats:
        summary_data = {
            'Categoría': ['Total Alertas', 'Alertas Recientes (24h)', 'Alertas Revisadas'],
            'Cantidad': [stats['total_alerts'], stats['recent_alerts'], stats['reviewed_alerts']]
        }
        dataframes['Resumen'] = pd.DataFrame(summary_data)
    
    # --- CONTENIDO ESPECÍFICO SEGÚN EL TIPO DE REPORTE ---
    
    # REPORTE TIPO 1: ANÁLISIS DE CÁMARAS
    if report_type == 'cameras-analysis':
        # Datos de distribución por cámara
        if charts_data and 'by_camera' in charts_data:
            camera_data = []
            for i, camera in enumerate(charts_data['by_camera']['labels']):
                camera_data.append({
                    'Cámara': camera,
                    'Cantidad': charts_data['by_camera']['data'][i]
                })
            dataframes['DistribuciónCámaras'] = pd.DataFrame(camera_data)
        
        # Patrones horarios si están disponibles
        if charts_data and 'hourly_pattern' in charts_data:
            hourly_data = []
            for i, hour in enumerate(charts_data['hourly_pattern']['labels']):
                hourly_data.append({
                    'Hora': hour,
                    'Cantidad': charts_data['hourly_pattern']['data'][i]
                })
            dataframes['PatronesHorarios'] = pd.DataFrame(hourly_data)
        
        # Filtrar alertas para mostrar solo las relacionadas con cámaras
        if alerts_data:
            camera_alerts = []
            for alert in alerts_data:
                camera_alerts.append({
                    'Cámara': alert['nombre_posicion'],
                    'Fecha': alert['fecha_evento'],
                    'Hora': alert['hora_evento'],
                    'Tipo': alert.get('tipo_evento', 'Alerta'),
                    'Revisado': 'Sí' if alert['revisado'] else 'No'
                })
            if camera_alerts:
                dataframes['AlertasCámaras'] = pd.DataFrame(camera_alerts)
        
        # Generar CSV con datos específicos de cámaras
        combined_df = pd.DataFrame()
        
        # Agregar sección de información
        if 'Información' in dataframes:
            combined_df = pd.concat([combined_df, dataframes['Información'], pd.DataFrame([[''] * len(dataframes['Información'].columns)])], axis=0)
        
        # Agregar distribución por cámara
        if 'DistribuciónCámaras' in dataframes:
            combined_df = pd.concat([combined_df, 
                                    pd.DataFrame([['DISTRIBUCIÓN POR CÁMARA'] + [''] * (len(dataframes['DistribuciónCámaras'].columns) - 1)]),
                                    dataframes['DistribuciónCámaras'], 
                                    pd.DataFrame([[''] * len(dataframes['DistribuciónCámaras'].columns)])], axis=0)
        
        # Agregar patrones horarios
        if 'PatronesHorarios' in dataframes:
            combined_df = pd.concat([combined_df, 
                                    pd.DataFrame([['PATRONES HORARIOS'] + [''] * (len(dataframes['PatronesHorarios'].columns) - 1)]),
                                    dataframes['PatronesHorarios'], 
                                    pd.DataFrame([[''] * len(dataframes['PatronesHorarios'].columns)])], axis=0)
        
        # Agregar alertas de cámaras
        if 'AlertasCámaras' in dataframes:
            combined_df = pd.concat([combined_df, 
                                    pd.DataFrame([['ALERTAS DE CÁMARAS'] + [''] * (len(dataframes['AlertasCámaras'].columns) - 1)]),
                                    dataframes['AlertasCámaras']], axis=0)
        
        combined_df.to_csv(file_path, index=False)
        
    # REPORTE TIPO 2: ANÁLISIS DE SENSORES Y DURACIÓN
    elif report_type == 'sensors-duration':
        # Datos de duración por categoría
        if stats and 'duracion_categorias' in stats:
            duracion_data = {
                'Categoría': ['Bajo (<5s)', 'Moderado (5-20s)', 'Alto (20-60s)', 'Crítico (>60s)'],
                'Cantidad': [
                    stats['duracion_categorias']['bajo'],
                    stats['duracion_categorias']['moderado'],
                    stats['duracion_categorias']['alto'],
                    stats['duracion_categorias']['critico']
                ]
            }
            dataframes['CategoríasDuración'] = pd.DataFrame(duracion_data)
        
        # Duración promedio por cámara/sensor
        if stats and 'duracion_promedio' in stats:
            duracion_promedio_data = []
            for cam in stats['duracion_promedio']:
                duracion_promedio_data.append({
                    'Sensor/Cámara': cam['nombre_posicion'],
                    'Duración Promedio (s)': cam['duracion_promedio']
                })
            if duracion_promedio_data:
                dataframes['DuraciónPromedio'] = pd.DataFrame(duracion_promedio_data)
        
        # Eventos de larga duración
        if stats and 'eventos_largos' in stats:
            eventos_largos_data = []
            for evento in stats['eventos_largos']:
                eventos_largos_data.append({
                    'Sensor/Cámara': evento['nombre_posicion'],
                    'Fecha': evento['fecha_evento'],
                    'Hora': evento['hora_evento'],
                    'Duración (s)': evento['duracion_segundos'],
                    'Revisado': 'Sí' if evento['revisado'] else 'No'
                })
            if eventos_largos_data:
                dataframes['EventosLargos'] = pd.DataFrame(eventos_largos_data)
        
        # Generar CSV con datos específicos de sensores y duración
        combined_df = pd.DataFrame()
        
        # Agregar sección de información
        if 'Información' in dataframes:
            combined_df = pd.concat([combined_df, dataframes['Información'], pd.DataFrame([[''] * len(dataframes['Información'].columns)])], axis=0)
        
        # Agregar resumen
        if 'Resumen' in dataframes:
            combined_df = pd.concat([combined_df, 
                                    pd.DataFrame([['RESUMEN DE SENSORES'] + [''] * (len(dataframes['Resumen'].columns) - 1)]),
                                    dataframes['Resumen'], 
                                    pd.DataFrame([[''] * len(dataframes['Resumen'].columns)])], axis=0)
        
        # Agregar categorías de duración
        if 'CategoríasDuración' in dataframes:
            combined_df = pd.concat([combined_df, 
                                    pd.DataFrame([['CATEGORÍAS POR DURACIÓN'] + [''] * (len(dataframes['CategoríasDuración'].columns) - 1)]),
                                    dataframes['CategoríasDuración'], 
                                    pd.DataFrame([[''] * len(dataframes['CategoríasDuración'].columns)])], axis=0)
        
        # Agregar duración promedio
        if 'DuraciónPromedio' in dataframes:
            combined_df = pd.concat([combined_df, 
                                    pd.DataFrame([['DURACIÓN PROMEDIO POR SENSOR'] + [''] * (len(dataframes['DuraciónPromedio'].columns) - 1)]),
                                    dataframes['DuraciónPromedio'], 
                                    pd.DataFrame([[''] * len(dataframes['DuraciónPromedio'].columns)])], axis=0)
        
        # Agregar eventos largos
        if 'EventosLargos' in dataframes:
            combined_df = pd.concat([combined_df, 
                                    pd.DataFrame([['EVENTOS DE LARGA DURACIÓN'] + [''] * (len(dataframes['EventosLargos'].columns) - 1)]),
                                    dataframes['EventosLargos']], axis=0)
        
        combined_df.to_csv(file_path, index=False)
        
    # REPORTE TIPO 3: COMPLETO INTEGRADO (O CUALQUIER OTRO TIPO)
    else:  # report_type == 'complete-integrated' o cualquier otro
        # Categorías por duración
        if stats and 'duracion_categorias' in stats:
            duracion_data = {
                'Categoría': ['Bajo (<5s)', 'Moderado (5-20s)', 'Alto (20-60s)', 'Crítico (>60s)'],
                'Cantidad': [
                    stats['duracion_categorias']['bajo'],
                    stats['duracion_categorias']['moderado'],
                    stats['duracion_categorias']['alto'],
                    stats['duracion_categorias']['critico']
                ]
            }
            dataframes['Duración'] = pd.DataFrame(duracion_data)
        
        # Eventos de larga duración
        if stats and 'eventos_largos' in stats:
            eventos_largos_data = []
            for evento in stats['eventos_largos']:
                eventos_largos_data.append({
                    'Cámara': evento['nombre_posicion'],
                    'Fecha': evento['fecha_evento'],
                    'Hora': evento['hora_evento'],
                    'Duración (s)': evento['duracion_segundos'],
                    'Revisado': 'Sí' if evento['revisado'] else 'No'
                })
            if eventos_largos_data:
                dataframes['EventosLargos'] = pd.DataFrame(eventos_largos_data)
        
        # Distribución por día
        if charts_data and 'by_day' in charts_data:
            day_data = []
            for i, day in enumerate(charts_data['by_day']['labels']):
                day_data.append({
                    'Día': day,
                    'Cantidad': charts_data['by_day']['data'][i]
                })
            dataframes['PorDía'] = pd.DataFrame(day_data)
        
        # Distribución por cámara
        if charts_data and 'by_camera' in charts_data:
            camera_data = []
            for i, camera in enumerate(charts_data['by_camera']['labels']):
                camera_data.append({
                    'Cámara': camera,
                    'Cantidad': charts_data['by_camera']['data'][i]
                })
            dataframes['PorCámara'] = pd.DataFrame(camera_data)
        
        # Todas las alertas
        if alerts_data:
            alerts_df_data = []
            for alert in alerts_data:
                alerts_df_data.append({
                    'Cámara': alert['nombre_posicion'],
                    'Descripción': alert.get('descripcion', 'N/A'),
                    'Fecha': alert['fecha_evento'],
                    'Hora': alert['hora_evento'],
                    'Duración': alert.get('duracion_segundos', 'N/A'),
                    'Revisado': 'Sí' if alert['revisado'] else 'No'
                })
            if alerts_df_data:
                dataframes['Alertas'] = pd.DataFrame(alerts_df_data)
        
        # Generar CSV con datos integrados
        combined_df = pd.DataFrame()
        
        # Agregar sección de información
        if 'Información' in dataframes:
            combined_df = pd.concat([combined_df, dataframes['Información'], pd.DataFrame([[''] * len(dataframes['Información'].columns)])], axis=0)
        
        # Agregar sección de resumen
        if 'Resumen' in dataframes:
            combined_df = pd.concat([combined_df, 
                                    pd.DataFrame([['RESUMEN DE ALERTAS'] + [''] * (len(dataframes['Resumen'].columns) - 1)]),
                                    dataframes['Resumen'], 
                                    pd.DataFrame([[''] * len(dataframes['Resumen'].columns)])], axis=0)
        
        # Agregar sección de duración
        if 'Duración' in dataframes:
            combined_df = pd.concat([combined_df, 
                                    pd.DataFrame([['CATEGORÍAS POR DURACIÓN'] + [''] * (len(dataframes['Duración'].columns) - 1)]),
                                    dataframes['Duración'], 
                                    pd.DataFrame([[''] * len(dataframes['Duración'].columns)])], axis=0)
        
        # Agregar sección de eventos largos
        if 'EventosLargos' in dataframes:
            combined_df = pd.concat([combined_df, 
                                    pd.DataFrame([['EVENTOS MÁS LARGOS'] + [''] * (len(dataframes['EventosLargos'].columns) - 1)]),
                                    dataframes['EventosLargos'], 
                                    pd.DataFrame([[''] * len(dataframes['EventosLargos'].columns)])], axis=0)
        
        # Agregar sección de distribución por día
        if 'PorDía' in dataframes:
            combined_df = pd.concat([combined_df, 
                                    pd.DataFrame([['DISTRIBUCIÓN POR DÍA'] + [''] * (len(dataframes['PorDía'].columns) - 1)]),
                                    dataframes['PorDía'], 
                                    pd.DataFrame([[''] * len(dataframes['PorDía'].columns)])], axis=0)
        
        # Agregar sección de eventos por cámara
        if 'PorCámara' in dataframes:
            combined_df = pd.concat([combined_df, 
                                    pd.DataFrame([['EVENTOS POR CÁMARA'] + [''] * (len(dataframes['PorCámara'].columns) - 1)]),
                                    dataframes['PorCámara'], 
                                    pd.DataFrame([[''] * len(dataframes['PorCámara'].columns)])], axis=0)
        
        # Agregar sección de alertas
        if 'Alertas' in dataframes:
            combined_df = pd.concat([combined_df, 
                                    pd.DataFrame([['LISTADO DE ALERTAS'] + [''] * (len(dataframes['Alertas'].columns) - 1)]),
                                    dataframes['Alertas']], axis=0)
        
        combined_df.to_csv(file_path, index=False)
        
        # También generar archivos CSV individuales si el reporte es completo
        for sheet_name, df in dataframes.items():
            individual_file_path = get_unique_filename(os.path.join(reports_dir, f"{filename}_{sheet_name}"), 'csv')
            df.to_csv(individual_file_path, index=False)
            
            # También guardar en Excel
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Guardar el Excel
        writer.close()
    
    return file_path

def get_date_range_text(date_range, start_date, end_date):
    """Genera texto descriptivo para el rango de fechas del reporte"""
    if date_range == 'day':
        return 'Últimas 24 horas'
    elif date_range == 'week':
        return 'Últimos 7 días'
    elif date_range == 'month':
        return 'Últimos 30 días'
    elif date_range == 'custom' and start_date and end_date:
        return f'Desde {start_date} hasta {end_date}'
    else:
        return 'Últimos 7 días (predeterminado)'


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')